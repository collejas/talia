"""Servicios específicos para WhatsApp via Twilio."""

from __future__ import annotations

import asyncio
import io
import json
import re
import time
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Mapping
from uuid import UUID, uuid4

import httpx
from fastapi import HTTPException
from openpyxl import load_workbook
from openai import AsyncOpenAI
from xml.etree import ElementTree as ET

from app.assistants.manager import AssistantConfig
from app.assistants.runtime import (
    build_prompt_payload,
    filter_assistant_tools,
    resolve_assistant_spec,
)
from app.assistants.tool_runtime import (
    ToolRuntimeContext,
    classify_runtime_error,
    run_tool_loop,
)
from app.channels.whatsapp import tools as whatsapp_tools
from app.channels.whatsapp.routing import resolve_whatsapp_organizacion
from app.channels.whatsapp.routing import resolve_whatsapp_organizacion_by_phone_number_id
from app.core.config import settings
from app.core.logging import get_logger, log_event
from app.repositories.crm import CRMRepository, CRMRepositoryError
from app.services import conversation_summary, leads_geo, storage
from app.services import assistant_document_delivery
from app.services import openai as openai_service
from app.services import openai_usage_ledger
from app.services.prospeccion_whatsapp_atribucion import resolve_first_matching_rule
from app.services.context_formatter import build_crm_context_lines, build_location_context_lines
from app.services import tenant_runtime
from app.services import twilio as twilio_service
from app.services.metrics import metrics
from app.services.prospeccion_progress import progress_hub
from app.services.storage import StorageError
from app.channels.booking_context import build_booking_context_message
from app.services.catalog_context import (
    build_catalog_context,
    build_catalog_inventory_context,
    is_location_request,
    should_autoload_inventory_context,
)
from app.services.prospeccion_auto_promoter import auto_promote_prospecto
from app.services.assistant_reply_guard import evaluate_reply_quality
from app.services.time_utils import get_current_time_reference
from app.services.high_demand_mode import high_demand_controller
from app.services.phone_utils import normalize_phone

from . import schemas

logger = get_logger("app.channels.whatsapp")

DEFAULT_FALLBACK = (
    "Tu mensaje quedó registrado, pero tuve un problema momentáneo al responder. "
    "Intentemos nuevamente en unos instantes."
)

_BOOKING_CONFIRMATION_HINTS: tuple[str, ...] = (
    "confirmo tu cita",
    "confirmarte tu cita",
    "te confirmo tu cita",
    "cita confirmada",
    "visita confirmada",
    "tu visita está agendada",
    "tu visita esta agendada",
    "cita agendada",
    "visita agendada",
    "quedó agendada",
    "quedo agendada",
    "te espero el",
    "te esperamos el",
    "ya tengo todo preparado para tu cita",
    "queda todo listo",
    "quedo todo listo",
    "tu cita esta lista",
    "tu cita está lista",
    "quedo pendiente para tu visita",
    "quedó pendiente para tu visita",
    "visita mañana a las",
    "visita para mañana a las",
)

_DETAILED_REPLY_HINTS: tuple[str, ...] = (
    "detalles",
    "ficha",
    "caracteristicas",
    "características",
    "completo",
    "completa",
    "toda la info",
    "toda la información",
    "todas las opciones",
    "lista",
    "comparar",
    "comparación",
    "cotización",
    "cotizacion",
)
_DEFAULT_WHATSAPP_MAX_CHARS = 500
_MAX_PROSPECCION_REPLY_PREVIEW_CHARS = 500
_PROSPECCION_REPLY_ENVIO_SOURCE_STATES = {"pendiente", "procesando", "enviado", "entregado", "leido", "completado"}
_WHATSAPP_ATTRIB_CONTACT_DEDUP_MINUTES = 60 * 24
_WHATSAPP_TYPING_INDICATOR_URL = "https://messaging.twilio.com/v2/Indicators/Typing.json"
_WHATSAPP_READ_INDICATOR_URL = "https://messaging.twilio.com/v2/Indicators/Read.json"
_WHATSAPP_TYPING_TIMEOUT_SECONDS = 6.0
_WHATSAPP_READ_TIMEOUT_SECONDS = 6.0
_WHATSAPP_INBOUND_DEBOUNCE_SECONDS = 1.2
_WHATSAPP_INBOUND_MERGE_MAX_MESSAGES = 4
_WHATSAPP_INBOUND_MERGE_MAX_WINDOW_SECONDS = 12
_WHATSAPP_INBOUND_FRAGMENT_MAX_CHARS = 80
_WHATSAPP_INBOUND_FRAGMENT_MAX_WORDS = 12
MAX_WHATSAPP_ATTACHMENT_BYTES = 8 * 1024 * 1024
MAX_WHATSAPP_ATTACHMENT_TEXT_CHARS = 12000
MAX_WHATSAPP_ATTACHMENTS_PER_MESSAGE = 4
_MARKDOWN_LINK_RE = re.compile(r"\[([^\]]+)\]\((https?://[^)]+)\)")


def _persona_datos(persona: Mapping[str, Any] | None) -> dict[str, Any]:
    if not persona:
        return {}
    payload = persona.get("persona_datos")
    if isinstance(payload, dict):
        return dict(payload)
    payload = persona.get("contacto_datos")
    if isinstance(payload, dict):
        return dict(payload)
    payload = persona.get("metadata")
    if isinstance(payload, dict):
        return dict(payload)
    return {}


def _persona_legacy_contact_id(persona: Mapping[str, Any] | None) -> str | None:
    if not persona:
        return None
    persona_data = _persona_datos(persona)
    legacy_contact_id = str(persona_data.get("legacy_contacto_id") or "").strip()
    if not legacy_contact_id:
        return None
    try:
        return str(UUID(legacy_contact_id))
    except (TypeError, ValueError):
        return None


def _contact_email_value(contact: Mapping[str, Any] | None) -> str | None:
    if not contact:
        return None
    for key in ("correo_principal", "correo_secundario", "correo", "email"):
        value = contact.get(key)
        if isinstance(value, str):
            trimmed = value.strip()
            if trimmed:
                return trimmed
    return None


def _contact_phone_value(contact: Mapping[str, Any] | None) -> str | None:
    if not contact:
        return None
    for key in (
        "telefono_principal_e164",
        "telefono_movil_1_e164",
        "telefono_e164",
        "telefono",
        "telefono_secundario_e164",
        "telefono_movil_2_e164",
    ):
        value = contact.get(key)
        if isinstance(value, str):
            trimmed = value.strip()
            if trimmed:
                return trimmed
    return None


def _normalize_inbound_message_id(value: Any) -> str | None:
    parsed = str(value or "").strip()
    return parsed or None


def _log_trace_stage(
    *,
    stage: str,
    conversation_id: str | None,
    persona_id: str | None,
    inbound_message_id: str | None,
    extra: dict[str, Any] | None = None,
) -> None:
    payload: dict[str, Any] = {
        "stage": stage,
        "conversation_id": conversation_id,
        "persona_id": persona_id,
        "inbound_message_id": inbound_message_id,
    }
    if extra:
        payload.update(extra)
    log_event(logger, "whatsapp.message_trace", **payload)


def _record_stage_timing(
    stage_timings: dict[str, float],
    stage_name: str,
    started_at: float,
) -> None:
    stage_timings[stage_name] = round((time.perf_counter() - started_at) * 1000, 2)


def _log_turn_timing(
    *,
    trace_id: str,
    conversation_id: str | None,
    persona_id: str | None,
    inbound_message_id: str | None,
    total_started_at: float,
    stage_timings: dict[str, float],
    extra: dict[str, Any] | None = None,
) -> None:
    payload: dict[str, Any] = {
        "trace_id": trace_id,
        "conversation_id": conversation_id,
        "persona_id": persona_id,
        "inbound_message_id": inbound_message_id,
        "total_ms": round((time.perf_counter() - total_started_at) * 1000, 2),
        "stage_timings": dict(stage_timings),
    }
    if extra:
        payload.update(extra)
    log_event(logger, "whatsapp.turn_timing", **payload)


@dataclass(slots=True)
class AssistantReply:
    """Respuesta del asistente junto con metadatos para persistencia."""

    text: str | None
    openai_conversation_id: str | None
    response_id: str | None
    tools_called: list[str] | None = None
    debug_timings: dict[str, float] | None = None


@dataclass(slots=True)
class TwilioSendResult:
    """Resultado resumido del envío a través de Twilio."""

    sid: str | None
    status: str | None
    error: str | None = None
    from_number: str | None = None
    provider: str = "twilio"


def _trim_text(value: Any) -> str | None:
    if isinstance(value, str):
        trimmed = value.strip()
        return trimmed or None
    if value is None:
        return None
    trimmed = str(value).strip()
    return trimmed or None


def _build_whatsapp_openai_metadata_payload(
    *,
    conversation_id: str,
    persona_id: str,
    message_sid: str | None,
    inbound_message_id: str | None,
    prospeccion_mode: bool,
    origin_type: str | None,
) -> dict[str, str]:
    return {
        "conversation_id": conversation_id,
        "persona_id": persona_id,
        "channel": "whatsapp",
        "message_sid": str(message_sid or "").strip(),
        "inbound_message_id": str(inbound_message_id or "").strip(),
        "prospeccion_mode": str(bool(prospeccion_mode)).lower(),
        "origin_type": str(origin_type or "").strip().lower() or "general_whatsapp",
    }


def _summarize_openai_response_payload(payload: Mapping[str, Any] | None) -> dict[str, Any]:
    if not isinstance(payload, Mapping):
        return {}
    output_items = payload.get("output") or []
    output_types: list[str] = []
    message_text_count = 0
    for item in output_items:
        if not isinstance(item, Mapping):
            continue
        item_type = str(item.get("type") or "").strip()
        if item_type:
            output_types.append(item_type)
        for content in item.get("content") or []:
            if not isinstance(content, Mapping):
                continue
            if str(content.get("type") or "").strip() == "output_text":
                text = content.get("text")
                if isinstance(text, str) and text.strip():
                    message_text_count += 1
    return {
        "response_id": _trim_text(payload.get("id")),
        "status": _trim_text(payload.get("status")),
        "output_count": len(output_items) if isinstance(output_items, list) else None,
        "output_types": output_types[:6],
        "output_text_count": message_text_count,
    }


def _normalize_whatsapp_provider(value: Any) -> str:
    provider = str(value or "").strip().lower()
    return provider if provider in {"twilio", "meta"} else "twilio"


def _conversation_inbox_context(conversation_meta: Mapping[str, Any] | None) -> dict[str, Any]:
    if not isinstance(conversation_meta, Mapping):
        return {}
    raw = conversation_meta.get("inbox_context")
    if isinstance(raw, dict):
        return dict(raw)
    if isinstance(raw, str):
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError:
            return {}
        if isinstance(parsed, dict):
            return dict(parsed)
    return {}


def _should_auto_send_welcome_document(
    *,
    whatsapp_settings: tenant_runtime.WhatsappRuntimeSettings,
    conversation_meta: Mapping[str, Any] | None,
) -> bool:
    welcome_prompt_version = str(
        whatsapp_settings.welcome_document_prompt_version or ""
    ).strip()
    if not welcome_prompt_version:
        return False
    if str(whatsapp_settings.prompt_version or "").strip() != welcome_prompt_version:
        return False
    inbox_context = _conversation_inbox_context(conversation_meta)
    if bool(inbox_context.get("welcome_document_sent")):
        return False
    last_response_id = str((conversation_meta or {}).get("last_response_id") or "").strip()
    if last_response_id:
        return False
    return True


async def _send_welcome_document_for_conversation(
    *,
    conversation_id: str,
    persona_id: str,
    message: schemas.WhatsAppIncomingMessage,
    whatsapp_settings: tenant_runtime.WhatsappRuntimeSettings,
    organizacion_id: UUID | None,
    conversation_meta: Mapping[str, Any] | None,
) -> bool:
    if not _should_auto_send_welcome_document(
        whatsapp_settings=whatsapp_settings,
        conversation_meta=conversation_meta,
    ):
        return False

    context = ToolRuntimeContext(
        conversation_id=conversation_id,
        persona_id=persona_id,
        channel="whatsapp",
        organizacion_id=str(organizacion_id) if organizacion_id else None,
    )
    try:
        documents = await assistant_document_delivery.resolve_documents_for_context(
            context=context,
            channel_scope="whatsapp",
            category="welcome",
            limit=1,
        )
    except Exception as exc:
        logger.warning(
            "whatsapp.welcome_documents_lookup_failed",
            extra={"conversation_id": conversation_id, "error": str(exc)},
        )
        return False

    if not documents:
        try:
            documents = await assistant_document_delivery.resolve_documents_for_context(
                context=context,
                channel_scope="whatsapp",
                category=None,
                limit=1,
            )
        except Exception as exc:
            logger.warning(
                "whatsapp.welcome_documents_fallback_lookup_failed",
                extra={"conversation_id": conversation_id, "error": str(exc)},
            )
            return False

    if not documents:
        return False

    attachments = assistant_document_delivery.build_whatsapp_attachments(documents)
    if not attachments:
        return False

    try:
        send_result = await send_manual_message(
            to_number=message.from_number,
            body=None,
            attachments=[attachments[0]],
            organizacion_id=organizacion_id,
        )
    except Exception as exc:
        logger.warning(
            "whatsapp.welcome_document_send_failed",
            extra={"conversation_id": conversation_id, "error": str(exc)},
        )
        return False
    if send_result.status != "sent" or not send_result.sid:
        logger.warning(
            "whatsapp.welcome_document_send_not_sent",
            extra={
                "conversation_id": conversation_id,
                "status": send_result.status,
                "error": send_result.error,
                "provider": send_result.provider,
            },
        )
        return False

    document = documents[0]
    try:
        await storage.merge_conversation_inbox_context(
            conversation_id,
            {
                "welcome_document_sent": True,
                "welcome_document_id": str(document.get("id") or "").strip() or None,
                "welcome_document_title": str(document.get("title") or "").strip() or None,
                "welcome_document_channel": "whatsapp",
                "welcome_document_sent_at": datetime.now(timezone.utc).isoformat(),
                "welcome_document_provider_sid": str(send_result.sid or "").strip() or None,
            },
        )
    except StorageError as exc:
        logger.warning(
            "whatsapp.welcome_document_context_mark_failed",
            extra={"conversation_id": conversation_id, "error": str(exc)},
        )
    return True


def _guess_attachment_name(attachment: schemas.WhatsAppMediaAttachment) -> str:
    candidate = attachment.filename or attachment.provider_id or attachment.url or ""
    candidate = str(candidate).strip()
    if candidate:
        return Path(candidate).name
    return f"whatsapp-adjunto-{attachment.index + 1}"


def _guess_attachment_extension(name: str, content_type: str | None) -> str:
    suffix = Path(name).suffix.lower()
    if suffix:
        return suffix
    mime = (content_type or "").lower().strip()
    mime_map = {
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/png": ".png",
        "image/webp": ".webp",
        "image/gif": ".gif",
        "audio/ogg": ".ogg",
        "audio/mpeg": ".mp3",
        "audio/mp3": ".mp3",
        "audio/wav": ".wav",
        "audio/x-wav": ".wav",
        "video/mp4": ".mp4",
        "application/pdf": ".pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
        "application/msword": ".doc",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
        "application/vnd.ms-excel": ".xls",
    }
    return mime_map.get(mime, "")


def _extract_pdf_attachments_from_reply_text(text: str) -> tuple[str, list[dict[str, Any]]]:
    """Convierte enlaces markdown a PDFs en adjuntos reales para WhatsApp."""

    if not text:
        return text, []

    attachments: list[dict[str, Any]] = []
    cleaned_text = text
    for match in _MARKDOWN_LINK_RE.finditer(text):
        label = _trim_text(match.group(1)) or "documento"
        url = _trim_text(match.group(2)) or ""
        if not url:
            continue
        if ".pdf" not in url.lower():
            continue
        filename = Path(label).name
        if not filename.lower().endswith(".pdf"):
            filename = f"{filename}.pdf"
        attachments.append(
            {
                "url": url,
                "mime": "application/pdf",
                "name": filename,
            }
        )
        cleaned_text = cleaned_text.replace(match.group(0), "").strip()

    if attachments:
        cleaned_text = re.sub(r"\s{2,}", " ", cleaned_text).strip()
    return cleaned_text, attachments[:MAX_WHATSAPP_ATTACHMENTS_PER_MESSAGE]


def _is_image_mime(mime: str | None, *, extension: str | None = None) -> bool:
    mime_value = (mime or "").lower()
    ext_value = (extension or "").lower()
    return mime_value.startswith("image/") or ext_value in {".jpg", ".jpeg", ".png", ".webp", ".gif"}


def _is_audio_mime(mime: str | None, *, extension: str | None = None) -> bool:
    mime_value = (mime or "").lower()
    ext_value = (extension or "").lower()
    return mime_value.startswith("audio/") or ext_value in {".ogg", ".opus", ".mp3", ".wav", ".aac", ".amr", ".3gp"}


def _is_text_attachment(mime: str | None, *, extension: str | None = None) -> bool:
    mime_value = (mime or "").lower()
    ext_value = (extension or "").lower()
    return mime_value.startswith("text/") or ext_value in {".txt", ".csv", ".md", ".json", ".log"}


def _is_docx_attachment(mime: str | None, *, extension: str | None = None) -> bool:
    mime_value = (mime or "").lower()
    ext_value = (extension or "").lower()
    return ext_value == ".docx" or mime_value == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"


def _is_xlsx_attachment(mime: str | None, *, extension: str | None = None) -> bool:
    mime_value = (mime or "").lower()
    ext_value = (extension or "").lower()
    return ext_value in {".xlsx", ".xlsm"} or mime_value in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroenabled.12",
    }


def _is_pdf_attachment(mime: str | None, *, extension: str | None = None) -> bool:
    mime_value = (mime or "").lower()
    ext_value = (extension or "").lower()
    return ext_value == ".pdf" or mime_value == "application/pdf"


def _trim_text_payload(text: str, limit: int) -> tuple[str, bool]:
    normalized = str(text or "").strip()
    if len(normalized) <= limit:
        return normalized, False
    return normalized[:limit].rstrip(), True


def _extract_docx_text(data: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        xml_bytes = archive.read("word/document.xml")
    root = ET.fromstring(xml_bytes)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    for para in root.findall(".//w:p", ns):
        runs = [node.text for node in para.findall(".//w:t", ns) if node.text]
        if runs:
            paragraphs.append("".join(runs))
    return "\n\n".join(paragraphs).strip()


def _extract_xlsx_text(data: bytes) -> str:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    try:
        for sheet in workbook.worksheets:
            sheet_lines: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if values:
                    sheet_lines.append(" | ".join(values))
            if sheet_lines:
                lines.append(f"[{sheet.title}]")
                lines.extend(sheet_lines)
    finally:
        workbook.close()
    return "\n".join(lines).strip()


async def _transcribe_audio_attachment(
    client: AsyncOpenAI,
    *,
    data: bytes,
    filename: str,
    content_type: str | None,
    model_name: str | None,
) -> str | None:
    model = model_name or settings.openai_stt_model or "whisper-1"
    try:
        response = await client.audio.transcriptions.create(
            model=model,
            file=(filename, io.BytesIO(data), content_type or "application/octet-stream"),
        )
    except Exception as exc:  # pragma: no cover - depende del SDK/red
        logger.warning(
            "whatsapp.attachment_transcription_failed",
            extra={"name": filename, "error": str(exc)},
        )
        return None
    text = getattr(response, "text", None)
    if isinstance(text, str) and text.strip():
        return text.strip()
    if isinstance(response, str) and response.strip():
        return response.strip()
    return None


async def _download_whatsapp_attachment(
    *,
    attachment: schemas.WhatsAppMediaAttachment,
    provider: str,
    whatsapp_settings: tenant_runtime.WhatsappRuntimeSettings,
) -> tuple[bytes | None, str | None, str]:
    download_url = attachment.url
    resolved_mime = attachment.content_type
    resolved_name = _guess_attachment_name(attachment)

    if provider == "meta" and attachment.provider_id:
        meta_token = getattr(whatsapp_settings, "meta_page_access_token", None)
        if not meta_token:
            return None, resolved_mime, resolved_name
        graph_version = getattr(whatsapp_settings, "meta_graph_api_version", None) or "v21.0"
        media_url = f"https://graph.facebook.com/{graph_version}/{attachment.provider_id}"
        headers = {
            "Authorization": f"Bearer {meta_token}",
            "Accept": "application/json",
        }
        async with httpx.AsyncClient(timeout=10.0, follow_redirects=True) as client:
            response = await client.get(media_url, headers=headers)
        if response.status_code >= 400:
            logger.warning(
                "whatsapp.meta_media_lookup_failed",
                extra={
                    "provider_id": attachment.provider_id,
                    "status_code": response.status_code,
                    "body": response.text[:300],
                },
            )
            return None, resolved_mime, resolved_name
        payload = response.json() if response.headers.get("content-type", "").lower().startswith("application/json") else {}
        if isinstance(payload, dict):
            download_url = payload.get("url") if isinstance(payload.get("url"), str) else download_url
            resolved_mime = resolved_mime or (
                payload.get("mime_type") if isinstance(payload.get("mime_type"), str) else None
            )
            if not attachment.filename:
                filename = payload.get("filename") or payload.get("name")
                if isinstance(filename, str) and filename.strip():
                    resolved_name = Path(filename).name
        if not download_url:
            return None, resolved_mime, resolved_name
        async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as client:
            response = await client.get(
                download_url,
                headers={"Authorization": f"Bearer {meta_token}"},
            )
        if response.status_code >= 400:
            logger.warning(
                "whatsapp.meta_media_download_failed",
                extra={
                    "provider_id": attachment.provider_id,
                    "status_code": response.status_code,
                    "body": response.text[:300],
                },
            )
            return None, resolved_mime, resolved_name
        return response.content, resolved_mime, resolved_name

    if not download_url:
        return None, resolved_mime, resolved_name

    twilio_account_sid = getattr(whatsapp_settings, "twilio_account_sid", None)
    twilio_auth_token = getattr(whatsapp_settings, "twilio_auth_token", None)
    async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as client:
        response = await client.get(download_url)
        if response.status_code in {401, 403} and twilio_account_sid and twilio_auth_token:
            response = await client.get(download_url, auth=(twilio_account_sid, twilio_auth_token))
    if response.status_code >= 400:
        logger.warning(
            "whatsapp.attachment_download_failed",
            extra={
                "url": download_url,
                "status_code": response.status_code,
                "body": response.text[:300],
            },
        )
        return None, resolved_mime, resolved_name
    return response.content, resolved_mime, resolved_name


async def _prepare_whatsapp_attachments(
    *,
    message: schemas.WhatsAppIncomingMessage,
    whatsapp_settings: tenant_runtime.WhatsappRuntimeSettings,
    client: AsyncOpenAI,
    conversation_id: str | None,
    provider: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    normalized_attachments: list[dict[str, Any]] = []
    content_items: list[dict[str, Any]] = []
    warnings: list[str] = []

    if not message.media:
        return normalized_attachments, content_items, warnings

    processed = 0
    for attachment in message.media:
        if processed >= MAX_WHATSAPP_ATTACHMENTS_PER_MESSAGE:
            warnings.append(
                f"Se ignoraron adjuntos extra; límite {MAX_WHATSAPP_ATTACHMENTS_PER_MESSAGE} por mensaje."
            )
            break

        data_bytes, mime, filename = await _download_whatsapp_attachment(
            attachment=attachment,
            provider=provider,
            whatsapp_settings=whatsapp_settings,
        )
        if not data_bytes:
            warnings.append(f"No se pudo descargar {filename}.")
            continue
        if len(data_bytes) > MAX_WHATSAPP_ATTACHMENT_BYTES:
            warnings.append(
                f"El archivo {filename} supera el límite permitido de {MAX_WHATSAPP_ATTACHMENT_BYTES // (1024 * 1024)} MB."
            )
            continue

        extension = _guess_attachment_extension(filename, mime)
        try:
            uploaded = await storage.upload_whatsapp_attachment(
                content=data_bytes,
                filename=filename,
                content_type=mime,
                conversation_id=conversation_id,
            )
        except StorageError as exc:
            logger.warning(
                "whatsapp.attachment_upload_failed",
                extra={"name": filename, "error": str(exc)},
            )
            warnings.append(f"No se pudo guardar {filename} en el almacenamiento.")
            continue

        normalized = dict(uploaded)
        if attachment.provider_id and not normalized.get("provider_id"):
            normalized["provider_id"] = attachment.provider_id
        normalized_attachments.append(normalized)

        if _is_image_mime(mime, extension=extension):
            content_items.append({"type": "input_image", "image_url": uploaded["url"]})
            processed += 1
            continue

        if _is_audio_mime(mime, extension=extension):
            transcript = await _transcribe_audio_attachment(
                client,
                data=data_bytes,
                filename=filename,
                content_type=mime,
                model_name=getattr(whatsapp_settings, "voice_stt_model", None),
            )
            if transcript:
                trimmed_text, truncated = _trim_text_payload(transcript, MAX_WHATSAPP_ATTACHMENT_TEXT_CHARS)
                if truncated:
                    trimmed_text += (
                        f"\n\n[Nota interna: contenido truncado a {MAX_WHATSAPP_ATTACHMENT_TEXT_CHARS} caracteres.]"
                    )
                content_items.append(
                    {
                        "type": "input_text",
                        "text": f"Transcripción de {filename}:\n{trimmed_text}",
                    }
                )
            else:
                warnings.append(f"No pude transcribir {filename}.")
            processed += 1
            continue

        if _is_docx_attachment(mime, extension=extension):
            try:
                doc_text = _extract_docx_text(data_bytes)
            except Exception as exc:  # pragma: no cover - archivos corruptos
                logger.warning(
                    "whatsapp.attachment_docx_parse_failed",
                    extra={"name": filename, "error": str(exc)},
                )
                warnings.append(f"No pude leer {filename} (DOCX).")
                continue
            trimmed_text, truncated = _trim_text_payload(doc_text, MAX_WHATSAPP_ATTACHMENT_TEXT_CHARS)
            if truncated:
                trimmed_text += (
                    f"\n\n[Nota interna: contenido truncado a {MAX_WHATSAPP_ATTACHMENT_TEXT_CHARS} caracteres.]"
                )
            content_items.append(
                {
                    "type": "input_text",
                    "text": f"Contenido de {filename} (extraído de DOCX):\n{trimmed_text}",
                }
            )
            processed += 1
            continue

        if _is_xlsx_attachment(mime, extension=extension):
            try:
                sheet_text = _extract_xlsx_text(data_bytes)
            except Exception as exc:  # pragma: no cover - archivos corruptos
                logger.warning(
                    "whatsapp.attachment_xlsx_parse_failed",
                    extra={"name": filename, "error": str(exc)},
                )
                warnings.append(f"No pude leer {filename} (XLSX).")
                continue
            trimmed_text, truncated = _trim_text_payload(sheet_text, MAX_WHATSAPP_ATTACHMENT_TEXT_CHARS)
            if truncated:
                trimmed_text += (
                    f"\n\n[Nota interna: contenido truncado a {MAX_WHATSAPP_ATTACHMENT_TEXT_CHARS} caracteres.]"
                )
            content_items.append(
                {
                    "type": "input_text",
                    "text": f"Contenido de {filename} (extraído de Excel):\n{trimmed_text}",
                }
            )
            processed += 1
            continue

        if _is_pdf_attachment(mime, extension=extension):
            try:
                upload = await client.files.create(
                    file=(filename, io.BytesIO(data_bytes), mime or "application/pdf"),
                    purpose="assistants",
                )
            except Exception as exc:  # pragma: no cover - API/red
                logger.warning(
                    "whatsapp.attachment_pdf_upload_failed",
                    extra={"name": filename, "error": str(exc)},
                )
                warnings.append(f"No se pudo compartir {filename} con el asistente.")
                continue
            content_items.append({"type": "input_file", "file_id": upload.id})
            processed += 1
            continue

        if _is_text_attachment(mime, extension=extension):
            try:
                text_content = data_bytes.decode("utf-8")
            except UnicodeDecodeError:
                text_content = data_bytes.decode("utf-8", errors="replace")
            trimmed_text, truncated = _trim_text_payload(text_content, MAX_WHATSAPP_ATTACHMENT_TEXT_CHARS)
            if truncated:
                trimmed_text += (
                    f"\n\n[Nota interna: contenido truncado a {MAX_WHATSAPP_ATTACHMENT_TEXT_CHARS} caracteres.]"
                )
            content_items.append(
                {
                    "type": "input_text",
                    "text": f"Contenido de {filename}:\n{trimmed_text}",
                }
            )
            processed += 1
            continue

        try:
            upload = await client.files.create(
                file=(filename, io.BytesIO(data_bytes), mime or "application/octet-stream"),
                purpose="assistants",
            )
        except Exception as exc:  # pragma: no cover - API/red
            logger.warning(
                "whatsapp.attachment_upload_failed",
                extra={"name": filename, "error": str(exc)},
            )
            warnings.append(f"No se pudo compartir {filename} con el asistente.")
            continue

        content_items.append({"type": "input_file", "file_id": upload.id})
        processed += 1

    return normalized_attachments, content_items, warnings


def _normalize_meta_template_language(value: Any) -> str | None:
    language = str(value or "").strip()
    return language or None


def _normalize_meta_template_name(value: Any) -> str | None:
    name = str(value or "").strip()
    return name or None


def _build_meta_template_components(
    content_variables: dict[str, str] | None,
) -> list[dict[str, Any]] | None:
    if not content_variables:
        return None

    def _sort_key(item: tuple[str, str]) -> tuple[int, int | str]:
        key = str(item[0]).strip()
        if key.isdigit():
            return (0, int(key))
        return (1, key)

    parameters: list[dict[str, Any]] = []
    for _, value in sorted(content_variables.items(), key=_sort_key):
        text_value = str(value or "")
        parameters.append({"type": "text", "text": text_value})
    if not parameters:
        return None
    return [{"type": "body", "parameters": parameters}]


def _build_meta_template_payload(
    *,
    template_name: str,
    template_language: str,
    content_variables: dict[str, str] | None = None,
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "type": "template",
        "template": {
            "name": template_name,
            "language": {
                "code": template_language,
                "policy": "deterministic",
            },
        },
    }
    components = _build_meta_template_components(content_variables)
    if components:
        payload["template"]["components"] = components
    return payload


def _normalize_meta_recipient_number(value: str | None) -> str | None:
    if not value:
        return None
    cleaned = value.strip()
    if cleaned.lower().startswith("whatsapp:"):
        cleaned = cleaned.split(":", 1)[1]
    digits = re.sub(r"\D+", "", cleaned)
    if digits:
        return digits
    return cleaned or None


async def _sync_inbound_to_prospeccion_log(
    *,
    repo: CRMRepository,
    conversation_id: str | None = None,
    organizacion_id: UUID | None = None,
    persona_id: str,
    message: schemas.WhatsAppIncomingMessage,
) -> bool:
    """Registra respuesta entrante en bitácora de prospección cuando aplica."""

    try:
        contact_uuid = UUID(persona_id)
    except (TypeError, ValueError):
        return False

    try:
        prospecto = await repo.worker_find_prospecto_by_contacto(
            contacto_id=contact_uuid,
            organizacion_id=organizacion_id,
        )
    except CRMRepositoryError as exc:
        log_event(
            logger,
            "whatsapp.prospeccion_reply_lookup_failed",
            persona_id=persona_id,
            error=str(exc),
        )
        return False
    envio: dict[str, Any] | None = None
    phone_candidates = _phone_lookup_candidates(message.from_number)
    if not prospecto:
        latest_envios_by_phone: dict[str, dict[str, Any]] = {}
        if phone_candidates:
            try:
                latest_envios_by_phone = await repo.worker_get_latest_envios_by_phones(
                    phone_values=set(phone_candidates),
                    organizacion_id=organizacion_id,
                    canal="whatsapp",
                )
            except CRMRepositoryError as exc:
                log_event(
                    logger,
                    "whatsapp.prospeccion_reply_envio_phone_lookup_failed",
                    persona_id=persona_id,
                    phone=",".join(phone_candidates),
                    error=str(exc),
                )
        for phone_candidate in phone_candidates:
            envio = latest_envios_by_phone.get(phone_candidate)
            if not envio:
                continue
            envio_prospecto_id = envio.get("prospecto_id")
            if not envio_prospecto_id:
                continue
            prospecto = {"id": envio_prospecto_id}
            log_event(
                logger,
                "whatsapp.prospeccion_reply_context_resolved_by_phone",
                persona_id=persona_id,
                phone=phone_candidate,
                prospecto_id=str(envio_prospecto_id),
                envio_id=str(envio.get("id")) if envio.get("id") else None,
            )
            break
    if not prospecto:
        latest_prospectos_by_phone: dict[str, dict[str, Any]] = {}
        if phone_candidates:
            try:
                latest_prospectos_by_phone = await repo.worker_get_latest_prospectos_by_phones(
                    phone_values=set(phone_candidates),
                    organizacion_id=organizacion_id,
                )
            except CRMRepositoryError as exc:
                log_event(
                    logger,
                    "whatsapp.prospeccion_reply_prospect_phone_lookup_failed",
                    persona_id=persona_id,
                    phone=",".join(phone_candidates),
                    error=str(exc),
                )
        for phone_candidate in phone_candidates:
            prospecto = latest_prospectos_by_phone.get(phone_candidate)
            if not prospecto:
                continue
            log_event(
                logger,
                "whatsapp.prospeccion_reply_context_resolved_by_prospect_phone",
                persona_id=persona_id,
                phone=phone_candidate,
                prospecto_id=str(prospecto.get("id")) if prospecto.get("id") else None,
            )
            break
    if not prospecto:
        return False

    prospecto_id = prospecto.get("id")
    if not prospecto_id:
        return False
    try:
        prospecto_uuid = UUID(str(prospecto_id))
    except (TypeError, ValueError):
        return False

    if envio is None:
        try:
            envio = await repo.worker_get_latest_envio_for_prospecto(
                prospecto_id=prospecto_uuid,
                canal="whatsapp",
            )
        except CRMRepositoryError as exc:
            log_event(
                logger,
                "whatsapp.prospeccion_reply_envio_lookup_failed",
                prospecto_id=str(prospecto_id),
                error=str(exc),
            )

    envio_id_value = envio.get("id") if envio else None
    batch_id_value = envio.get("batch_id") if envio else None
    body_text = _trim_text(message.body)
    if body_text and len(body_text) > _MAX_PROSPECCION_REPLY_PREVIEW_CHARS:
        body_text = f"{body_text[:_MAX_PROSPECCION_REPLY_PREVIEW_CHARS]}..."

    if envio_id_value:
        try:
            envio_uuid = UUID(str(envio_id_value))
        except (TypeError, ValueError):
            envio_uuid = None
        if envio_uuid:
            envio_estado = _trim_text(envio.get("estado")) if isinstance(envio, dict) else None
            envio_estado_norm = (envio_estado or "").lower()
            if envio_estado_norm in _PROSPECCION_REPLY_ENVIO_SOURCE_STATES:
                current_detalle = envio.get("detalle") if isinstance(envio.get("detalle"), dict) else {}
                update_payload = {
                    "estado": "respondido",
                    "detalle": {
                        **current_detalle,
                        "reply_inbound_at": datetime.now(timezone.utc).isoformat(),
                        "reply_message_sid": _trim_text(message.message_sid),
                        "reply_preview": body_text,
                    },
                    "error": None,
                    "procesado_en": datetime.now(timezone.utc).isoformat(),
                }
                try:
                    await repo.worker_complete_envio(envio_id=envio_uuid, payload=update_payload)
                except CRMRepositoryError as exc:
                    log_event(
                        logger,
                        "whatsapp.prospeccion_reply_envio_update_failed",
                        envio_id=str(envio_uuid),
                        error=str(exc),
                    )

    if conversation_id:
        inbox_context_patch = _build_prospeccion_inbox_context_patch(envio)
        try:
            await storage.merge_conversation_inbox_context(conversation_id, inbox_context_patch)
        except StorageError as exc:
            log_event(
                logger,
                "whatsapp.prospeccion_reply_inbox_context_failed",
                conversation_id=conversation_id,
                prospecto_id=str(prospecto_uuid),
                error=str(exc),
            )

    log_entry: dict[str, Any] = {
        "prospecto_id": str(prospecto_uuid),
        "canal": "whatsapp",
        "accion": "reply_inbound",
        "estado": "respondido",
        "detalle": {
            "action": "reply_inbound",
            "direction": "incoming",
            "respondio": True,
            "message_sid": _trim_text(message.message_sid),
            "wa_id": _trim_text(message.wa_id),
            "from_number": _trim_text(message.from_number),
            "body": body_text,
        },
    }
    organizacion_id_value = None
    if envio and envio.get("organizacion_id"):
        organizacion_id_value = envio.get("organizacion_id")
    elif isinstance(prospecto, dict) and prospecto.get("organizacion_id"):
        organizacion_id_value = prospecto.get("organizacion_id")
    elif organizacion_id:
        organizacion_id_value = str(organizacion_id)
    if organizacion_id_value:
        log_entry["organizacion_id"] = str(organizacion_id_value)
    if envio_id_value:
        log_entry["envio_id"] = str(envio_id_value)
    if batch_id_value:
        log_entry["batch_id"] = str(batch_id_value)
    try:
        await repo.worker_insert_contact_logs([log_entry])
    except CRMRepositoryError as exc:
        log_event(
            logger,
            "whatsapp.prospeccion_reply_log_failed",
            prospecto_id=str(prospecto_uuid),
            envio_id=str(envio_id_value) if envio_id_value else None,
            error=str(exc),
        )
        return True

    if batch_id_value:
        try:
            await progress_hub.publish(
                str(batch_id_value),
                {
                    "type": "reply",
                    "batch_id": str(batch_id_value),
                    "envio_id": str(envio_id_value) if envio_id_value else None,
                    "prospecto_id": str(prospecto_uuid),
                    "canal": "whatsapp",
                },
            )
            if envio_id_value:
                await progress_hub.publish(
                    str(batch_id_value),
                    {
                        "type": "envio",
                        "batch_id": str(batch_id_value),
                        "envio_id": str(envio_id_value),
                        "estado": "respondido",
                    },
                )
        except Exception as exc:  # pragma: no cover - SSE best effort
            log_event(
                logger,
                "whatsapp.prospeccion_reply_publish_failed",
                batch_id=str(batch_id_value),
                prospecto_id=str(prospecto_uuid),
                error=str(exc),
            )
    return True


async def _resolve_prospeccion_prospecto_id(
    *,
    repo: CRMRepository,
    persona_id: str,
    organizacion_id: UUID | None,
    message: schemas.WhatsAppIncomingMessage,
) -> UUID | None:
    """Resuelve prospecto relacionado al inbound usando contacto o teléfono."""
    try:
        contact_uuid = UUID(persona_id)
    except (TypeError, ValueError):
        contact_uuid = None
    if contact_uuid:
        try:
            by_contact = await repo.worker_find_prospecto_by_contacto(
                contacto_id=contact_uuid,
                organizacion_id=organizacion_id,
            )
        except CRMRepositoryError:
            by_contact = None
        prospecto_id = (by_contact or {}).get("id") if isinstance(by_contact, dict) else None
        try:
            return UUID(str(prospecto_id))
        except (TypeError, ValueError):
            pass
    phone_candidates = _phone_lookup_candidates(message.from_number)
    if not phone_candidates:
        return None
    try:
        latest_prospectos_by_phone = await repo.worker_get_latest_prospectos_by_phones(
            phone_values=set(phone_candidates),
            organizacion_id=organizacion_id,
        )
    except CRMRepositoryError:
        latest_prospectos_by_phone = {}
    for phone_candidate in phone_candidates:
        by_phone = latest_prospectos_by_phone.get(phone_candidate)
        prospecto_id = (by_phone or {}).get("id") if isinstance(by_phone, dict) else None
        try:
            return UUID(str(prospecto_id))
        except (TypeError, ValueError):
            continue
    return None


async def _maybe_apply_publicidad_whatsapp_attribution(
    *,
    repo: CRMRepository,
    organizacion_id: UUID,
    conversation_id: str,
    persona_id: str,
    message_id: str | None,
    message: schemas.WhatsAppIncomingMessage,
) -> dict[str, Any] | None:
    """Evalúa reglas de atribución publicitaria y persiste el primer match útil."""

    phrase_original = _trim_text(message.body)
    if not phrase_original:
        return None

    try:
        contact_uuid = UUID(persona_id)
    except (TypeError, ValueError):
        return None

    try:
        active_rules = await repo.list_active_whatsapp_atribucion_reglas(organizacion_id=organizacion_id)
    except CRMRepositoryError as exc:
        log_event(
            logger,
            "whatsapp.publicidad_atribucion_rules_fetch_failed",
            conversation_id=conversation_id,
            persona_id=persona_id,
            error=str(exc),
        )
        return None
    if not active_rules:
        return None

    matched_rule, applied_match_type, normalized_phrase = resolve_first_matching_rule(
        incoming_text=phrase_original,
        rules=active_rules,
    )
    if not matched_rule:
        log_event(
            logger,
            "whatsapp.publicidad_atribucion_no_match",
            conversation_id=conversation_id,
            persona_id=persona_id,
            phrase=normalized_phrase,
        )
        return None

    since_iso = (datetime.now(timezone.utc) - timedelta(minutes=_WHATSAPP_ATTRIB_CONTACT_DEDUP_MINUTES)).isoformat()
    try:
        persona_row = await repo.get_persona_by_id(persona_id=persona_id)
    except CRMRepositoryError as exc:
        log_event(
            logger,
            "whatsapp.publicidad_atribucion_persona_fetch_failed",
            conversation_id=conversation_id,
            persona_id=persona_id,
            error=str(exc),
        )
        persona_row = None
    legacy_contact_id = _persona_legacy_contact_id(persona_row)
    try:
        recent_event = await repo.worker_get_recent_whatsapp_atribucion_event_for_persona(
            organizacion_id=organizacion_id,
            persona_id=contact_uuid,
            legacy_contact_id=UUID(legacy_contact_id) if legacy_contact_id else None,
            since_iso=since_iso,
        )
    except CRMRepositoryError as exc:
        log_event(
            logger,
            "whatsapp.publicidad_atribucion_recent_lookup_failed",
            conversation_id=conversation_id,
            persona_id=persona_id,
            error=str(exc),
        )
        recent_event = None
    if isinstance(recent_event, dict):
        recent_conversation_id = _trim_text(recent_event.get("conversacion_id"))
        if recent_conversation_id and recent_conversation_id != conversation_id:
            log_event(
                logger,
                "whatsapp.publicidad_atribucion_skipped_recent_contact_window",
                conversation_id=conversation_id,
                persona_id=persona_id,
                recent_conversation_id=recent_conversation_id,
                window_minutes=_WHATSAPP_ATTRIB_CONTACT_DEDUP_MINUTES,
            )
            return None

    event_payload = {
        "organizacion_id": str(organizacion_id),
        "regla_id": _trim_text(matched_rule.get("id")),
        "conversacion_id": conversation_id,
        "persona_id": persona_id,
        "contacto_id": legacy_contact_id,
        "mensaje_id": _trim_text(message.message_sid),
        "frase_original": phrase_original,
        "frase_normalizada": normalized_phrase,
        "tipo_match": applied_match_type,
        "canal_publicitario": _trim_text(matched_rule.get("canal_publicitario")),
        "campana_publicitaria": _trim_text(matched_rule.get("campana_publicitaria")),
        "adset": _trim_text(matched_rule.get("adset")),
        "anuncio": _trim_text(matched_rule.get("anuncio")),
        "metadata": {
            "source": "whatsapp_inbound",
            "rule_name": _trim_text(matched_rule.get("nombre_regla")),
            "rule_priority": matched_rule.get("prioridad"),
        },
    }
    event_payload = {k: v for k, v in event_payload.items() if v is not None}
    try:
        created_event = await repo.worker_create_whatsapp_atribucion_event(
            organizacion_id=organizacion_id,
            payload=event_payload,
        )
    except CRMRepositoryError as exc:
        log_event(
            logger,
            "whatsapp.publicidad_atribucion_event_create_failed",
            conversation_id=conversation_id,
            persona_id=persona_id,
            error=str(exc),
        )
        return None
    if not created_event:
        return None

    if message_id:
        try:
            existing_message = await repo.worker_get_message_by_id(
                organizacion_id=organizacion_id,
                message_id=message_id,
            )
        except CRMRepositoryError as exc:
            log_event(
                logger,
                "whatsapp.publicidad_atribucion_message_fetch_failed",
                conversation_id=conversation_id,
                message_id=message_id,
                error=str(exc),
            )
            existing_message = None
        if isinstance(existing_message, dict):
            current_datos = (
                existing_message.get("datos")
                if isinstance(existing_message.get("datos"), dict)
                else {}
            )
            patched_datos = {
                **current_datos,
                "source": "publicidad_whatsapp",
                "source_detail": {
                    "canal_publicitario": event_payload.get("canal_publicitario"),
                    "campana_publicitaria": event_payload.get("campana_publicitaria"),
                    "adset": event_payload.get("adset"),
                    "anuncio": event_payload.get("anuncio"),
                    "regla_id": _trim_text(created_event.get("regla_id")) or _trim_text(matched_rule.get("id")),
                    "regla_nombre": _trim_text(matched_rule.get("nombre_regla")),
                    "tipo_match": applied_match_type,
                    "frase_normalizada": normalized_phrase,
                    "atribuido_en": _trim_text(created_event.get("creado_en"))
                    or datetime.now(timezone.utc).isoformat(),
                },
            }
            try:
                await repo.worker_update_message_datos(
                    organizacion_id=organizacion_id,
                    message_id=message_id,
                    datos=patched_datos,
                )
            except CRMRepositoryError as exc:
                log_event(
                    logger,
                    "whatsapp.publicidad_atribucion_message_update_failed",
                    conversation_id=conversation_id,
                    message_id=message_id,
                    error=str(exc),
                )

    if isinstance(persona_row, dict):
        persona_data = _persona_datos(persona_row)
        patched_persona_data = {
            **persona_data,
            "publicidad_whatsapp_atribucion": {
                "source": "publicidad_whatsapp",
                "regla_id": _trim_text(created_event.get("regla_id")) or _trim_text(matched_rule.get("id")),
                "regla_nombre": _trim_text(matched_rule.get("nombre_regla")),
                "canal_publicitario": event_payload.get("canal_publicitario"),
                "campana_publicitaria": event_payload.get("campana_publicitaria"),
                "adset": event_payload.get("adset"),
                "anuncio": event_payload.get("anuncio"),
                "tipo_match": applied_match_type,
                "frase_normalizada": normalized_phrase,
                "conversacion_id": conversation_id,
                "atribuido_en": _trim_text(created_event.get("creado_en")) or datetime.now(timezone.utc).isoformat(),
            },
        }
        patch_payload: dict[str, Any] = {"persona_datos": patched_persona_data}
        current_origen = _trim_text(persona_row.get("origen"))
        if (current_origen or "").lower() in {"", "whatsapp", "prospeccion"}:
            patch_payload["origen"] = "publicidad_whatsapp"
        try:
            await repo.update_persona_by_id(persona_id=persona_id, patch=patch_payload)
        except CRMRepositoryError as exc:
            log_event(
                logger,
                "whatsapp.publicidad_atribucion_contact_update_failed",
                conversation_id=conversation_id,
                persona_id=persona_id,
                error=str(exc),
            )

    try:
        await storage.merge_conversation_inbox_context(
            conversation_id,
            {
                "source": "publicidad_whatsapp",
                "source_detail": {
                    "canal_publicitario": event_payload.get("canal_publicitario"),
                    "campana_publicitaria": event_payload.get("campana_publicitaria"),
                    "adset": event_payload.get("adset"),
                    "anuncio": event_payload.get("anuncio"),
                    "regla_id": _trim_text(created_event.get("regla_id")) or _trim_text(matched_rule.get("id")),
                    "regla_nombre": _trim_text(matched_rule.get("nombre_regla")),
                    "tipo_match": applied_match_type,
                    "frase_normalizada": normalized_phrase,
                    "atribuido_en": _trim_text(created_event.get("creado_en"))
                    or datetime.now(timezone.utc).isoformat(),
                },
            },
        )
    except StorageError as exc:
        log_event(
            logger,
            "whatsapp.publicidad_atribucion_conversation_snapshot_failed",
            conversation_id=conversation_id,
            persona_id=persona_id,
            error=str(exc),
        )

    log_event(
        logger,
        "whatsapp.publicidad_atribucion_applied",
        conversation_id=conversation_id,
        persona_id=persona_id,
        regla_id=_trim_text(matched_rule.get("id")),
        canal_publicitario=event_payload.get("canal_publicitario"),
        tipo_match=applied_match_type,
    )
    return created_event


async def _mark_opportunity_as_prospeccion_whatsapp(
    *,
    repo: CRMRepository,
    organizacion_id: UUID,
    opportunity_id: str | None,
    attribution_event: Mapping[str, Any] | None,
) -> None:
    if not opportunity_id or not attribution_event:
        return
    try:
        opportunity_uuid = UUID(str(opportunity_id))
    except (TypeError, ValueError):
        return
    try:
        opportunity = await repo.get_pipeline_opportunity(
            organizacion_id=organizacion_id,
            oportunidad_id=opportunity_uuid,
        )
    except CRMRepositoryError as exc:
        log_event(
            logger,
            "whatsapp.publicidad_atribucion_opportunity_fetch_failed",
            opportunity_id=opportunity_id,
            error=str(exc),
        )
        return
    if not isinstance(opportunity, dict):
        return
    metadata = opportunity.get("metadata")
    metadata_dict = dict(metadata) if isinstance(metadata, dict) else {}
    source_current = str(metadata_dict.get("source") or "").strip().lower()
    if source_current in {"", "assistant", "whatsapp", "publicidad_whatsapp", "whatsapp_inbound"}:
        metadata_dict["source"] = "prospeccion_whatsapp"
    metadata_dict["prospeccion_canal"] = "whatsapp"
    metadata_dict["publicidad_whatsapp_atribucion"] = {
        "source": "publicidad_whatsapp",
        "regla_id": str(attribution_event.get("regla_id") or "").strip() or None,
        "canal_publicitario": str(attribution_event.get("canal_publicitario") or "").strip() or None,
        "campana_publicitaria": str(attribution_event.get("campana_publicitaria") or "").strip() or None,
        "adset": str(attribution_event.get("adset") or "").strip() or None,
        "anuncio": str(attribution_event.get("anuncio") or "").strip() or None,
        "tipo_match": str(attribution_event.get("tipo_match") or "").strip() or None,
        "frase_normalizada": str(attribution_event.get("frase_normalizada") or "").strip() or None,
        "conversacion_id": str(attribution_event.get("conversacion_id") or "").strip() or None,
        "atribuido_en": str(attribution_event.get("creado_en") or "").strip() or datetime.now(timezone.utc).isoformat(),
    }
    try:
        await repo.update_opportunity(
            organizacion_id=organizacion_id,
            oportunidad_id=opportunity_uuid,
            payload={"metadata": metadata_dict},
        )
    except CRMRepositoryError as exc:
        log_event(
            logger,
            "whatsapp.publicidad_atribucion_opportunity_update_failed",
            opportunity_id=opportunity_id,
            error=str(exc),
        )


def _looks_like_booking_confirmation(text: str) -> bool:
    normalized = str(text or "").strip().lower()
    if not normalized:
        return False
    return any(hint in normalized for hint in _BOOKING_CONFIRMATION_HINTS)


def _wants_detailed_reply(text: str | None) -> bool:
    normalized = str(text or "").strip().lower()
    if not normalized:
        return False
    return any(hint in normalized for hint in _DETAILED_REPLY_HINTS)


def _compact_whatsapp_reply(text: str | None, max_chars: int = _DEFAULT_WHATSAPP_MAX_CHARS) -> str:
    compact = " ".join(str(text or "").split())
    if len(compact) <= max_chars:
        return compact
    return compact[: max_chars - 1].rstrip() + "…"


def _should_wait_for_inbound_burst(text: str | None) -> bool:
    normalized = " ".join(str(text or "").split())
    if not normalized:
        return False
    if any(punct in normalized for punct in ".!?"):
        return False
    return (
        len(normalized) <= _WHATSAPP_INBOUND_FRAGMENT_MAX_CHARS
        and len(normalized.split()) <= _WHATSAPP_INBOUND_FRAGMENT_MAX_WORDS
    )


def _is_unknown_prompt_variable_error(exc: Exception) -> bool:
    text = str(exc or "").lower()
    return "prompt_variable_unknown" in text or "unknown prompt variables" in text


def _is_previous_response_not_found_error(exc: Exception) -> bool:
    text = str(exc or "").lower()
    return "previous_response_not_found" in text or "previous response with id" in text


async def _refresh_conversation_summary_best_effort(
    *,
    conversation_id: str,
    persona_id: str | None,
    organizacion_id: UUID | None,
    context_data: dict[str, Any] | None,
) -> None:
    try:
        await conversation_summary.ensure_conversation_summary(
            conversation_id=conversation_id,
            persona_id=persona_id,
            context_data=context_data,
            organizacion_id=organizacion_id,
        )
    except Exception as exc:  # pragma: no cover
        logger.warning(
            "whatsapp.conversation_summary_refresh_failed",
            extra={"conversation_id": conversation_id, "error": str(exc)},
        )


async def _guard_booking_confirmation_claim(
    *,
    conversation_id: str,
    reply_text: str,
    persona: Mapping[str, Any] | None = None,
    opportunity_id: str | None = None,
) -> str:
    if not _looks_like_booking_confirmation(reply_text):
        return reply_text
    try:
        booking = await storage.fetch_calendar_booking_by_conversation(conversation_id)
    except StorageError as exc:
        logger.warning(
            "whatsapp.booking_guard_lookup_failed",
            extra={"conversation_id": conversation_id, "error": str(exc)},
        )
        return reply_text
    status = str((booking or {}).get("status") or "").strip().lower()
    if booking and status in {"confirmed", "reprogrammed"}:
        return reply_text
    log_event(
        logger,
        "whatsapp.booking_guard_blocked_false_confirmation",
        conversation_id=conversation_id,
        booking_status=status or None,
    )
    try:
        resolved_persona = dict(persona or {})
        resolved_opportunity_id = str(opportunity_id or "").strip() or None
        if (not resolved_persona) or not resolved_opportunity_id:
            conversation_meta = await storage.fetch_conversation(conversation_id)
            if not resolved_persona:
                persona_id_value = str(conversation_meta.get("persona_id") or "").strip()
                if persona_id_value:
                    resolved_persona = await storage.fetch_persona(persona_id_value)
            if not resolved_opportunity_id and resolved_persona:
                resolved_opportunity_id = await storage.ensure_persona_conversation_opportunity(
                    conversation_id=conversation_id,
                    persona_id=str(resolved_persona.get("id") or ""),
                    channel="whatsapp",
                )
        if resolved_persona and resolved_opportunity_id:
            prefilter_status = await whatsapp_tools._has_prefilter_for_schedule(
                contact=resolved_persona,
                opportunity_id=resolved_opportunity_id,
                conversation_id=conversation_id,
            )
            missing_fields = [
                str(item).strip()
                for item in (prefilter_status.get("missing_fields") or [])
                if str(item).strip()
            ]
            if missing_fields:
                questions = prefilter_status.get("questions") or {}
                if isinstance(questions, Mapping):
                    missing_field = missing_fields[0]
                    question_text = str(
                        questions.get(missing_field)
                        or ""
                    ).strip()
                    if question_text:
                        return f"Para confirmar tu cita, solo falta este dato: {question_text}"
    except Exception as exc:
        logger.warning(
            "whatsapp.booking_guard_prefilter_lookup_failed",
            extra={"conversation_id": conversation_id, "error": str(exc)},
        )
    return (
        "Para confirmar tu cita, aún falta un dato breve. "
        "Te hago una pregunta rápida para continuar."
    )


async def handle_incoming_message(
    message: schemas.WhatsAppIncomingMessage,
    source: str = "webhook",
    organizacion_id: UUID | str | None = None,
) -> None:
    """Procesa un mensaje entrante desde Twilio y delega la respuesta a OpenAI."""
    turn_started = time.perf_counter()
    trace_id = f"whatsapp-{uuid4().hex[:12]}"
    stage_timings: dict[str, float] = {}
    org_uuid: UUID | None = None
    if isinstance(organizacion_id, UUID):
        org_uuid = organizacion_id
    elif isinstance(organizacion_id, str):
        org_uuid = _parse_org_uuid(organizacion_id)
    log_event(
        logger,
        "whatsapp.incoming_message_received",
        message_sid=message.message_sid,
        source=source,
    )

    if await _maybe_handle_sales_acknowledgement(message):
        return

    if message.message_sid:
        try:
            duplicate_check_started = time.perf_counter()
            existing_message = await storage.fetch_message_by_twilio_sid(message.message_sid)
            _record_stage_timing(stage_timings, "duplicate_check_ms", duplicate_check_started)
        except StorageError as exc:
            logger.warning(
                "whatsapp.fetch_message_by_sid_failed",
                extra={"message_sid": message.message_sid, "error": str(exc)},
            )
        else:
            if existing_message and (existing_message.get("direccion") == "entrante"):
                log_event(
                    logger,
                    "whatsapp.duplicate_incoming_ignored",
                    message_sid=message.message_sid,
                    source=source,
                )
                return

    normalized_from = _normalize_phone_number(message.from_number)
    recipient_number = _normalize_phone_number(message.to_number)
    if org_uuid is None:
        resolve_org_started = time.perf_counter()
        phone_number_id = getattr(message, "phone_number_id", None)
        if phone_number_id:
            organizacion_hint = await resolve_whatsapp_organizacion_by_phone_number_id(
                phone_number_id=phone_number_id
            )
        else:
            organizacion_hint = None
        if not organizacion_hint:
            organizacion_hint = await resolve_whatsapp_organizacion(to_number=recipient_number)
        _record_stage_timing(stage_timings, "resolve_org_ms", resolve_org_started)
        org_uuid = _parse_org_uuid(organizacion_hint) if organizacion_hint else None
    else:
        organizacion_hint = str(org_uuid)

    if not organizacion_hint:
        logger.error(
            "whatsapp.organizacion_unresolved",
            extra={"to_number": recipient_number, "wa_id": message.wa_id},
        )
        raise HTTPException(status_code=500, detail="No se pudo enrutar el mensaje entrante")

    runtime_settings_started = time.perf_counter()
    whatsapp_settings = await tenant_runtime.get_whatsapp_runtime_settings(
        organizacion_id=org_uuid,
    )
    _record_stage_timing(stage_timings, "runtime_settings_ms", runtime_settings_started)
    whatsapp_provider = _normalize_whatsapp_provider(whatsapp_settings.provider)
    normalized_attachments: list[dict[str, Any]] = []
    attachment_content_items: list[dict[str, Any]] = []
    attachment_warnings: list[str] = []
    if message.media:
        try:
            openai_client = openai_service.get_assistant_client(
                api_key=getattr(whatsapp_settings, "voice_api_key", None),
                project_id=getattr(whatsapp_settings, "project_id", None),
            )
            normalized_attachments, attachment_content_items, attachment_warnings = await _prepare_whatsapp_attachments(
                message=message,
                whatsapp_settings=whatsapp_settings,
                client=openai_client,
                conversation_id=message.message_sid,
                provider=whatsapp_provider,
            )
        except Exception as exc:  # pragma: no cover - fallback defensivo
            logger.warning(
                "whatsapp.attachment_prepare_failed",
                extra={"message_sid": message.message_sid, "error": str(exc)},
            )
    attachments_payload = normalized_attachments or message.attachments_as_dict()
    try:
        register_inbound_started = time.perf_counter()
        registration = await storage.register_whatsapp_message(
            direction="entrante",
            wa_id=message.wa_id,
            phone_e164=normalized_from,
            body=message.body,
            message_sid=message.message_sid,
            profile_name=message.profile_name,
            inactivity_minutes=whatsapp_settings.inactivity_minutes,
            metadata=message.metadata(),
            attachments=attachments_payload,
            webhook_payload=message.raw_payload,
            organizacion_id=organizacion_hint,
        )
        _record_stage_timing(stage_timings, "register_inbound_ms", register_inbound_started)
    except StorageError as exc:
        logger.exception(
            "whatsapp.register_incoming_failed",
            extra={
                "error": str(exc),
                "resolved_organizacion_id": organizacion_hint,
                "to_number": recipient_number,
                "from_number": normalized_from,
            },
        )
        raise HTTPException(status_code=502, detail="whatsapp_register_failed") from exc

    conversation_id = str(registration.get("conversation_id") or "")
    persona_id = str(registration.get("persona_id") or registration.get("contact_id") or "")
    current_message_id = str(registration.get("message_id") or "")
    inbound_message_id = _normalize_inbound_message_id(current_message_id)
    openai_conversation_id = registration.get("openai_conversation_id")
    org_uuid = _parse_org_uuid(organizacion_hint)

    _log_trace_stage(
        stage="inbound_persisted",
        conversation_id=conversation_id,
        persona_id=persona_id,
        inbound_message_id=inbound_message_id,
        extra={
            "message_sid": _trim_text(message.message_sid),
            "source": source,
        },
    )
    await high_demand_controller.record_inbound(channel="whatsapp")

    if conversation_id:
        try:
            await storage.update_conversation(conversation_id, {"estado": "abierta"})
        except StorageError as exc:
            logger.warning(
                "whatsapp.conversation_reopen_failed",
                extra={"conversation_id": conversation_id, "error": str(exc)},
            )

    if not conversation_id or not persona_id:
        logger.error(
            "whatsapp.registration_missing_ids",
            extra={"conversation_id": conversation_id},
        )
        return

    burst_merge_started = time.perf_counter()
    should_continue, merged_body, burst_timings = await _coalesce_inbound_burst(
        conversation_id=conversation_id,
        current_message_id=current_message_id,
        fallback_body=message.body,
    )
    _record_stage_timing(stage_timings, "burst_merge_ms", burst_merge_started)
    if burst_timings:
        stage_timings["burst_merge_substages"] = burst_timings  # type: ignore[assignment]
    if not should_continue:
        _log_turn_timing(
            trace_id=trace_id,
            conversation_id=conversation_id,
            persona_id=persona_id,
            inbound_message_id=inbound_message_id,
            total_started_at=turn_started,
            stage_timings=stage_timings,
            extra={"ended_early": "coalesced_burst", "source": source},
        )
        return
    message.body = merged_body

    declared_name = whatsapp_tools.lead_tools._extract_name_from_message_text(message.body)
    if declared_name and not whatsapp_tools.lead_tools._is_placeholder_full_name(declared_name):
        try:
            await storage.update_persona(persona_id, {"nombre_completo": declared_name})
            logger.info(
                "whatsapp.inbound_declared_name_applied",
                extra={
                    "conversation_id": conversation_id,
                    "persona_id": persona_id,
                    "declared_name": declared_name,
                },
            )
        except StorageError as exc:
            logger.warning(
                "whatsapp.inbound_declared_name_failed",
                extra={
                    "conversation_id": conversation_id,
                    "persona_id": persona_id,
                    "error": str(exc),
                },
            )

    try:
        repo = CRMRepository()
    except CRMRepositoryError as exc:
        log_event(logger, "whatsapp.prospeccion_reply_repo_error", error=str(exc))
        repo = None
    is_prospeccion_context = False
    publicidad_atribucion_event: Mapping[str, Any] | None = None
    if repo:
        prospeccion_sync_started = time.perf_counter()
        is_prospeccion_context = await _sync_inbound_to_prospeccion_log(
            repo=repo,
            conversation_id=conversation_id,
            organizacion_id=org_uuid,
            persona_id=persona_id,
            message=message,
        )
        publicidad_atribucion_event = await _maybe_apply_publicidad_whatsapp_attribution(
            repo=repo,
            organizacion_id=org_uuid,
            conversation_id=conversation_id,
            persona_id=persona_id,
            message_id=current_message_id,
            message=message,
        )
        _record_stage_timing(stage_timings, "prospeccion_sync_ms", prospeccion_sync_started)
    origin_type = (
        "publicidad_whatsapp"
        if publicidad_atribucion_event
        else "prospeccion"
        if is_prospeccion_context
        else "general_whatsapp"
    )
    is_prospeccion_mode = is_prospeccion_context or bool(publicidad_atribucion_event)
    if origin_type == "general_whatsapp":
        try:
            await storage.merge_conversation_inbox_context(
                conversation_id,
                {
                    "source": "general_whatsapp",
                    "source_detail": {"channel": "whatsapp", "mode": "assistant"},
                },
            )
        except StorageError as exc:
            logger.warning(
                "whatsapp.general_inbox_context_failed",
                extra={"conversation_id": conversation_id, "error": str(exc)},
            )

    restart_context: dict[str, Any] | None = None
    opportunity_ref: str | None = None
    ensure_persona_id = persona_id
    ensure_opportunity_started: float | None = None
    if repo:
        ensure_opportunity_started = time.perf_counter()
        prospecto_uuid = await _resolve_prospeccion_prospecto_id(
            repo=repo,
            persona_id=persona_id,
            organizacion_id=org_uuid,
            message=message,
        )
        if prospecto_uuid:
            try:
                prospecto_opportunity = await repo.worker_find_opportunity_by_prospecto(
                    organizacion_id=org_uuid,
                    prospecto_id=prospecto_uuid,
                )
            except CRMRepositoryError as exc:
                log_event(
                    logger,
                    "whatsapp.prospeccion_existing_opportunity_lookup_failed",
                    prospecto_id=str(prospecto_uuid),
                    error=str(exc),
                )
                prospecto_opportunity = None
            if isinstance(prospecto_opportunity, dict):
                opportunity_ref = str(prospecto_opportunity.get("id") or "").strip() or None
                prospect_persona_id = str(
                    prospecto_opportunity.get("contacto_principal_id") or ""
                ).strip()
                if prospect_persona_id:
                    ensure_persona_id = prospect_persona_id
                    log_event(
                        logger,
                        "whatsapp.prospeccion_reuse_opportunity_persona",
                        conversation_id=conversation_id,
                        prospecto_id=str(prospecto_uuid),
                        opportunity_id=opportunity_ref,
                        persona_id=prospect_persona_id,
                        inbound_persona_id=persona_id,
                    )
    try:
        ensure_payload = await storage.ensure_persona_conversation_opportunity(
            conversation_id=conversation_id,
            persona_id=ensure_persona_id,
            channel="whatsapp",
            include_restart_metadata=True,
        )
    except StorageError as exc:
        logger.warning(
            "whatsapp.ensure_opportunity_failed",
            extra={"conversation_id": conversation_id, "error": str(exc)},
        )
    else:
        if isinstance(ensure_payload, dict):
            restart_context = ensure_payload
            opportunity_ref = str(ensure_payload.get("oportunidad_id") or "").strip() or None
        else:
            opportunity_ref = str(ensure_payload or "").strip() or None
            restart_context = {
                "oportunidad_id": ensure_payload,
                "restart_created": False,
                "restart_sequence": 1,
            }
        if ensure_opportunity_started is not None:
            _record_stage_timing(stage_timings, "ensure_opportunity_ms", ensure_opportunity_started)
    if repo and opportunity_ref and publicidad_atribucion_event:
        await _mark_opportunity_as_prospeccion_whatsapp(
            repo=repo,
            organizacion_id=org_uuid,
            opportunity_id=opportunity_ref,
            attribution_event=publicidad_atribucion_event,
        )

    persona_record = await _maybe_update_persona_location(persona_id=persona_id)

    restart_created = bool(restart_context and restart_context.get("restart_created"))
    if restart_created:
        restart_sequence = int(restart_context.get("restart_sequence") or 1)
        opportunity_ref = str(restart_context.get("oportunidad_id") or opportunity_ref or "").strip() or None
        context = ToolRuntimeContext(
            conversation_id=conversation_id,
            persona_id=persona_id,
            channel="whatsapp",
        )
        resumen_text = f"El contacto retomó la conversación (ciclo #{restart_sequence})."
        notes_text = message.body or "El contacto reactivó la conversación."
        try:
            await whatsapp_tools._notify_sales_rep(
                context=context,
                trigger="restart_conversation",
                persona=persona_record,
                opportunity_id=opportunity_ref,
                resumen=resumen_text,
                notes=notes_text,
                email=None,
                extra={"restart_sequence": restart_sequence},
            )
        except Exception as exc:  # pragma: no cover - defensivo
            logger.warning(
                "whatsapp.restart_notify_failed",
                extra={"conversation_id": conversation_id, "error": str(exc)},
            )

    try:
        fetch_conversation_started = time.perf_counter()
        conversation_meta = await storage.fetch_conversation(conversation_id)
        _record_stage_timing(stage_timings, "fetch_conversation_ms", fetch_conversation_started)
    except StorageError as exc:
        logger.exception(
            "whatsapp.fetch_conversation_failed",
            extra={"conversation_id": conversation_id, "error": str(exc)},
        )
        return

    if conversation_meta.get("manual_override"):
        log_event(
            logger,
            "whatsapp.manual_override_active",
            conversation_id=conversation_id,
        )
        return

    previous_response_id = conversation_meta.get("last_response_id")
    if not openai_conversation_id:
        openai_conversation_id = conversation_meta.get("openai_conversation_id")

    catalog_inmobiliario_enabled = True
    catalog_no_inmobiliario_enabled = True
    if org_uuid:
        try:
            catalog_toggle_started = time.perf_counter()
            catalog_inmobiliario_enabled = await tenant_runtime.is_catalog_inmobiliario_enabled(
                organizacion_id=org_uuid,
                channel="whatsapp",
            )
            catalog_no_inmobiliario_enabled = await tenant_runtime.is_catalog_no_inmobiliario_enabled(
                organizacion_id=org_uuid,
                channel="whatsapp",
            )
            stage_timings["catalog_toggle_ms"] = round(
                (time.perf_counter() - catalog_toggle_started) * 1000, 2
            )
        except Exception as exc:  # pragma: no cover
            logger.warning(
                "whatsapp.catalog_toggle_lookup_failed",
                extra={
                    "conversation_id": conversation_id,
                    "organizacion_id": str(org_uuid),
                    "error": str(exc),
                },
            )
    inventory_context_text = None
    location_request = is_location_request(message.body or "")
    if catalog_inmobiliario_enabled and not location_request and should_autoload_inventory_context(message.body or ""):
        inventory_context_started = time.perf_counter()
        try:
            inventory_context_text = await build_catalog_inventory_context(organizacion_hint)
        except Exception as exc:
            logger.warning(
                "whatsapp.inventory_context_failed",
                extra={"conversation_id": conversation_id, "error": str(exc)},
            )
        _record_stage_timing(stage_timings, "catalog_inventory_context_ms", inventory_context_started)

    catalog_context = None
    if (catalog_inmobiliario_enabled or catalog_no_inmobiliario_enabled) and settings.catalog_context_autoload and not location_request:
        catalog_context_started = time.perf_counter()
        catalog_domain = "any"
        if catalog_inmobiliario_enabled and not catalog_no_inmobiliario_enabled:
            catalog_domain = "inmobiliario"
        elif catalog_no_inmobiliario_enabled and not catalog_inmobiliario_enabled:
            catalog_domain = "no_inmobiliario"
        catalog_context = await build_catalog_context(
            organizacion_hint,
            message.body or "",
            user_id=message.wa_id or message.from_number,
            channel="whatsapp",
            domain=catalog_domain,
        )
        _record_stage_timing(stage_timings, "catalog_context_ms", catalog_context_started)
    catalog_context_text = None
    if inventory_context_text and catalog_context:
        catalog_context_text = f"{inventory_context_text}\n\n{catalog_context.text}"
    elif inventory_context_text:
        catalog_context_text = inventory_context_text
    elif catalog_context:
        catalog_context_text = catalog_context.text
    booking_context_text = None
    try:
        booking_context_started = time.perf_counter()
        booking_context_text = await build_booking_context_message(
            persona_id=persona_id,
            conversation_id=conversation_id,
            channel="whatsapp",
            persona=persona_record,
        )
        _record_stage_timing(stage_timings, "booking_context_ms", booking_context_started)
    except Exception as exc:
        logger.warning(
            "whatsapp.booking_context_failed",
            extra={
                "conversation_id": conversation_id,
                "persona_id": persona_id,
                "error": str(exc),
            },
        )

    # Best-effort UX: marcar leído y mostrar "escribiendo..." mientras se procesa la respuesta.
    read_indicator_started = time.perf_counter()
    typing_indicator_started = time.perf_counter()
    await asyncio.gather(
        _send_whatsapp_read_indicator(
            incoming_message_sid=message.message_sid,
            organizacion_id=org_uuid,
        ),
        _send_whatsapp_typing_indicator(
            incoming_message_sid=message.message_sid,
            organizacion_id=org_uuid,
        ),
        return_exceptions=True,
    )
    _record_stage_timing(stage_timings, "read_indicator_ms", read_indicator_started)
    _record_stage_timing(stage_timings, "typing_indicator_ms", typing_indicator_started)

    try:
        assistant_generation_started = time.perf_counter()
        _log_trace_stage(
            stage="assistant_generation_started",
            conversation_id=conversation_id,
            persona_id=persona_id,
            inbound_message_id=inbound_message_id,
            extra={
                "previous_response_id": _trim_text(previous_response_id),
                "openai_conversation_id": _trim_text(openai_conversation_id),
            },
        )
        assistant_reply = await _generate_assistant_reply(
            message=message,
            conversation_id=conversation_id,
            persona_id=persona_id,
            openai_conversation_id=openai_conversation_id,
            previous_response_id=previous_response_id,
            catalog_context=catalog_context_text,
            booking_context=booking_context_text,
            whatsapp_settings=whatsapp_settings,
            organizacion_id=org_uuid,
            catalog_inmobiliario_enabled=catalog_inmobiliario_enabled,
            catalog_no_inmobiliario_enabled=catalog_no_inmobiliario_enabled,
            prospeccion_mode=is_prospeccion_mode,
            origin_type=origin_type,
            inbound_message_id=inbound_message_id,
            attachment_content_items=attachment_content_items
            + (
                [
                    {
                        "type": "input_text",
                        "text": (
                            "Nota interna: Algunos adjuntos no pudieron procesarse. "
                            + " ".join(attachment_warnings)
                        ).strip(),
                    }
                ]
                if attachment_warnings
                else []
            ),
        )
        _record_stage_timing(stage_timings, "assistant_generation_ms", assistant_generation_started)
        if assistant_reply.debug_timings:
            stage_timings["assistant_generation_substages"] = assistant_reply.debug_timings  # type: ignore[assignment]
        log_event(
            logger,
            "whatsapp.reply_generated",
            conversation_id=conversation_id,
            response_id=assistant_reply.response_id,
            openai_conversation_id=assistant_reply.openai_conversation_id,
            inbound_message_id=inbound_message_id,
        )
        _log_trace_stage(
            stage="assistant_generated",
            conversation_id=conversation_id,
            persona_id=persona_id,
            inbound_message_id=inbound_message_id,
            extra={
                "response_id": _trim_text(assistant_reply.response_id),
                "openai_conversation_id": _trim_text(assistant_reply.openai_conversation_id),
            },
        )
    except Exception as exc:  # pragma: no cover - errores inesperados de OpenAI
        error_meta = classify_runtime_error(exc)
        logger.exception(
            "whatsapp.generate_reply_failed",
            extra={
                "conversation_id": conversation_id,
                "error": str(exc),
                "error_type": error_meta.get("error_type"),
                "status_code": error_meta.get("status_code"),
                "retryable": bool(error_meta.get("retryable")),
                "inbound_message_id": inbound_message_id,
            },
        )
        assistant_reply = AssistantReply(
            text=DEFAULT_FALLBACK,
            openai_conversation_id=openai_conversation_id,
            response_id=previous_response_id,
            tools_called=[],
        )

    welcome_document_sent_by_tool = bool(
        assistant_reply.tools_called and "send_information_package" in assistant_reply.tools_called
    )

    if not assistant_reply.text:
        log_event(
            logger,
            "whatsapp.empty_reply",
            conversation_id=conversation_id,
        )
        empty_retry_text: str | None = None
        try:
            empty_retry_kwargs: dict[str, Any] = {
                "input": [
                    {
                        "role": "developer",
                        "content": [
                            {
                                "type": "input_text",
                                "text": (
                                    "Regenera SOLO un mensaje final de WhatsApp completo y autocontenido. "
                                    "Responde al último mensaje del cliente con 1-3 frases, máximo 300 caracteres. "
                                    "No uses herramientas, no expliques el error y no dejes la respuesta vacía."
                                ),
                            }
                        ],
                    }
                ],
                "store": True,
                "max_output_tokens": 180,
                "temperature": 0.2,
                "metadata": metadata_payload,
                "tool_choice": "none",
            }
            empty_retry_kwargs.update(_build_request_template(include_tools=False))
            if assistant.is_prompt:
                empty_retry_kwargs.pop("temperature", None)
            if assistant_reply.openai_conversation_id or openai_conversation_id:
                empty_retry_kwargs["conversation"] = (
                    assistant_reply.openai_conversation_id or openai_conversation_id
                )
            elif assistant_reply.response_id or previous_response_id:
                empty_retry_kwargs["previous_response_id"] = (
                    assistant_reply.response_id or previous_response_id
                )
            empty_retry_started = time.perf_counter()
            empty_retry_response = await client.responses.create(**empty_retry_kwargs)
            debug_timings["empty_retry_ms"] = round((time.perf_counter() - empty_retry_started) * 1000, 2)
            empty_retry_payload = empty_retry_response.model_dump()
            await openai_usage_ledger.record_response_usage(
                organizacion_id=organizacion_id,
                channel="whatsapp",
                feature="sales_chat",
                assistant=assistant,
                response_payload=empty_retry_payload,
                request_purpose="empty_reply_retry",
                latency_ms=int(round(debug_timings["empty_retry_ms"])),
                api_key=whatsapp_settings.voice_api_key,
                request_metadata={"conversation_id": conversation_id},
                conversation_id=conversation_id,
                persona_id=persona_id,
                fallback_used=True,
                project_id=assistant.project_id,
            )
            empty_retry_text = _extract_text_from_response(empty_retry_payload)
        except Exception as exc:
            logger.warning(
                "whatsapp.empty_reply_retry_failed",
                extra={"conversation_id": conversation_id, "error": str(exc)},
            )

        assistant_reply = AssistantReply(
            text=(
                empty_retry_text
                or (
                    "Ya te compartí el PDF de bienvenida. ¿Me compartes tu nombre y apellido, por favor?"
                    if welcome_document_sent_by_tool
                    else DEFAULT_FALLBACK
                )
            ),
            openai_conversation_id=assistant_reply.openai_conversation_id or openai_conversation_id,
            response_id=assistant_reply.response_id or previous_response_id,
            tools_called=assistant_reply.tools_called,
        )

    final_reply_text = assistant_reply.text
    if not final_reply_text:
        final_reply_text = DEFAULT_FALLBACK
    try:
        final_reply_text = await _guard_booking_confirmation_claim(
            conversation_id=conversation_id,
            reply_text=final_reply_text,
            persona=persona_record,
            opportunity_id=opportunity_ref,
        )
    except Exception as exc:
        logger.exception(
            "whatsapp.reply_guard_failed",
            extra={"conversation_id": conversation_id, "error": str(exc)},
        )
        log_event(
            logger,
            "whatsapp.reply_guard_failed",
            conversation_id=conversation_id,
            inbound_message_id=inbound_message_id,
            error=str(exc),
        )
    final_reply_text = _compact_whatsapp_reply(final_reply_text, _DEFAULT_WHATSAPP_MAX_CHARS)
    final_reply_text, reply_attachments = _extract_pdf_attachments_from_reply_text(final_reply_text)

    try:
        twilio_send_started = time.perf_counter()
        send_result = await _send_whatsapp_reply(
            to_number=message.from_number,
            body=final_reply_text,
            attachments=reply_attachments or None,
            organizacion_id=org_uuid,
        )
        _record_stage_timing(stage_timings, "twilio_send_ms", twilio_send_started)
    except asyncio.CancelledError:
        log_event(
            logger,
            "whatsapp.reply_dispatch_cancelled",
            conversation_id=conversation_id,
            inbound_message_id=inbound_message_id,
            stage="before_dispatched_log",
        )
        _log_trace_stage(
            stage="dispatch_cancelled",
            conversation_id=conversation_id,
            persona_id=persona_id,
            inbound_message_id=inbound_message_id,
        )
        _log_turn_timing(
            trace_id=trace_id,
            conversation_id=conversation_id,
            persona_id=persona_id,
            inbound_message_id=inbound_message_id,
            total_started_at=turn_started,
            stage_timings=stage_timings,
            extra={
                "source": source,
                "fallback_used": final_reply_text == DEFAULT_FALLBACK,
                "assistant_response_id": assistant_reply.response_id,
                "openai_conversation_id": assistant_reply.openai_conversation_id,
                "message_sid": _trim_text(message.message_sid),
                "delivery_status": "cancelled",
            },
        )
        raise
    except Exception as exc:
        error_meta = classify_runtime_error(exc)
        logger.exception(
            "whatsapp.reply_dispatch_failed",
            extra={
                "conversation_id": conversation_id,
                "inbound_message_id": inbound_message_id,
                "error": str(exc),
                "error_type": error_meta.get("error_type"),
                "status_code": error_meta.get("status_code"),
                "retryable": bool(error_meta.get("retryable")),
            },
        )
        log_event(
            logger,
            "whatsapp.send_skipped_unexpected",
            conversation_id=conversation_id,
            inbound_message_id=inbound_message_id,
            error=str(exc),
            error_type=error_meta.get("error_type"),
            status_code=error_meta.get("status_code"),
            retryable=bool(error_meta.get("retryable")),
        )
        if whatsapp_provider == "twilio":
            await high_demand_controller.record_twilio_attempt(error_code="dispatch_exception")
        await high_demand_controller.record_assistant_latency(
            channel="whatsapp",
            latency_ms=(time.perf_counter() - turn_started) * 1000,
        )
        _log_turn_timing(
            trace_id=trace_id,
            conversation_id=conversation_id,
            persona_id=persona_id,
            inbound_message_id=inbound_message_id,
            total_started_at=turn_started,
            stage_timings=stage_timings,
            extra={
                "source": source,
                "fallback_used": final_reply_text == DEFAULT_FALLBACK,
                "assistant_response_id": assistant_reply.response_id,
                "openai_conversation_id": assistant_reply.openai_conversation_id,
                "message_sid": _trim_text(message.message_sid),
                "delivery_status": "dispatch_failed",
            },
        )
        return

    log_event(
        logger,
        "whatsapp.reply_dispatched",
        conversation_id=conversation_id,
        status=send_result.status,
        error=send_result.error,
        inbound_message_id=inbound_message_id,
    )
    _log_trace_stage(
        stage="dispatch_attempted",
        conversation_id=conversation_id,
        persona_id=persona_id,
        inbound_message_id=inbound_message_id,
        extra={
            "provider": send_result.provider,
            "provider_message_id": _trim_text(send_result.sid),
            "delivery_status": _trim_text(send_result.status),
            "delivery_error": _trim_text(send_result.error),
        },
    )
    if send_result.provider == "twilio":
        await high_demand_controller.record_twilio_attempt(error_code=send_result.error)
    await high_demand_controller.record_assistant_latency(
        channel="whatsapp",
        latency_ms=(time.perf_counter() - turn_started) * 1000,
    )

    metadata = {
        "openai_conversation_id": assistant_reply.openai_conversation_id,
        "response_id": assistant_reply.response_id,
        "delivery_status": send_result.status,
        "provider": send_result.provider,
        "inbound_message_id": inbound_message_id,
        "inbound_message_sid": _trim_text(message.message_sid),
        "provider_message_id": _trim_text(send_result.sid),
    }
    if send_result.error:
        metadata["delivery_error"] = send_result.error

    resolved_persona_org = await resolve_whatsapp_organizacion(contact=persona_record) or organizacion_hint
    try:
        persist_outbound_started = time.perf_counter()
        outgoing_registration = await storage.register_whatsapp_message(
            direction="saliente",
            conversation_id=conversation_id,
            persona_id=persona_id,
            body=final_reply_text,
            message_sid=send_result.sid,
            response_id=assistant_reply.response_id,
            metadata=metadata,
            wa_id=message.wa_id,
            phone_e164=normalized_from,
            organizacion_id=resolved_persona_org,
        )
        _record_stage_timing(stage_timings, "persist_outbound_ms", persist_outbound_started)
    except StorageError as exc:
        logger.warning(
            "whatsapp.register_outgoing_failed",
            extra={"conversation_id": conversation_id, "error": str(exc)},
        )
        log_event(
            logger,
            "whatsapp.reply_register_failed",
            conversation_id=conversation_id,
            error=str(exc),
            inbound_message_id=inbound_message_id,
        )
    else:
        log_event(
            logger,
            "whatsapp.reply_registered",
            conversation_id=conversation_id,
            message_sid=send_result.sid,
            inbound_message_id=inbound_message_id,
            outbound_message_id=outgoing_registration.get("message_id"),
        )
        try:
            from app.services import whatsapp_followups as whatsapp_followup_jobs

            await whatsapp_followup_jobs.schedule_customer_followup(
                conversation_id=conversation_id,
                persona_id=persona_id,
                organizacion_id=str(resolved_persona_org) if resolved_persona_org else None,
                reason="assistant_reply",
            )
        except Exception as exc:
            logger.warning(
                "whatsapp.followup.schedule_failed",
                extra={"conversation_id": conversation_id, "error": str(exc)},
            )
        try:
            if welcome_document_sent_by_tool:
                log_event(
                    logger,
                    "whatsapp.welcome_document_skipped_after_tool",
                    conversation_id=conversation_id,
                    inbound_message_id=inbound_message_id,
                )
            else:
                welcome_sent = await _send_welcome_document_for_conversation(
                    conversation_id=conversation_id,
                    persona_id=persona_id,
                    message=message,
                    whatsapp_settings=whatsapp_settings,
                    organizacion_id=org_uuid,
                    conversation_meta=conversation_meta,
                )
                if welcome_sent:
                    log_event(
                        logger,
                        "whatsapp.welcome_document_sent",
                        conversation_id=conversation_id,
                        inbound_message_id=inbound_message_id,
                    )
        except Exception as exc:  # pragma: no cover - defensivo
            logger.warning(
                "whatsapp.welcome_document_flow_failed",
                extra={"conversation_id": conversation_id, "error": str(exc)},
            )
        _log_trace_stage(
            stage="assistant_persisted",
            conversation_id=conversation_id,
            persona_id=persona_id,
            inbound_message_id=inbound_message_id,
            extra={
                "outbound_message_id": outgoing_registration.get("message_id"),
                "provider": send_result.provider,
                "provider_message_id": _trim_text(send_result.sid),
            },
        )
    _log_turn_timing(
        trace_id=trace_id,
        conversation_id=conversation_id,
        persona_id=persona_id,
        inbound_message_id=inbound_message_id,
        total_started_at=turn_started,
        stage_timings=stage_timings,
        extra={
            "source": source,
            "fallback_used": final_reply_text == DEFAULT_FALLBACK,
            "assistant_response_id": assistant_reply.response_id,
            "openai_conversation_id": assistant_reply.openai_conversation_id,
            "message_sid": _trim_text(message.message_sid),
            "provider": send_result.provider,
            "provider_message_id": _trim_text(send_result.sid),
            "delivery_status": _trim_text(send_result.status),
        },
    )


async def handle_status_callback(
    callback: schemas.WhatsAppStatusCallback,
    *,
    provider: str = "twilio",
) -> None:
    """Persistencia básica de los eventos de entrega reportados por el provider."""
    event = _map_status_to_event(callback.status)
    if not event:
        log_event(
            logger,
            "whatsapp.status_ignored",
            message_sid=callback.message_sid,
            status=callback.status,
        )
        return

    try:
        await storage.record_delivery_event(
            provider=_normalize_whatsapp_provider(provider),
            message_sid=callback.message_sid,
            event=event,
            raw_payload=callback.raw_payload,
            error_code=callback.error_code,
            provider_timestamp=callback.timestamp,
        )
    except StorageError as exc:
        logger.warning(
            "whatsapp.delivery_event_failed",
            extra={"message_sid": callback.message_sid, "error": str(exc)},
        )
    else:
        log_event(
            logger,
            "whatsapp.delivery_event_recorded",
            message_sid=callback.message_sid,
            event=event,
        )

    if event == "fallido":
        await _retry_failed_sales_notification(
            message_sid=callback.message_sid,
            error_code=callback.error_code,
        )

    await _sync_envio_status_from_whatsapp(callback)


async def _retry_failed_sales_notification(
    *,
    message_sid: str,
    error_code: str | None = None,
) -> None:
    sid = str(message_sid or "").strip()
    if not sid:
        return
    try:
        repo = CRMRepository()
    except CRMRepositoryError as exc:
        log_event(logger, "whatsapp.retry_notification.repo_error", message_sid=sid, error=str(exc))
        return
    try:
        assignment = await repo.get_sales_assignment_by_notification_sid(notification_sid=sid)
    except CRMRepositoryError as exc:
        log_event(
            logger,
            "whatsapp.retry_notification.lookup_failed",
            message_sid=sid,
            error=str(exc),
        )
        return
    if not assignment:
        return

    trigger_event = str(assignment.get("trigger_event") or "").strip().lower()
    if not trigger_event.startswith("notify_"):
        return
    trigger = trigger_event.removeprefix("notify_")

    assignment_metadata = assignment.get("metadata")
    assignment_metadata = assignment_metadata if isinstance(assignment_metadata, dict) else {}
    notification_meta = assignment_metadata.get("notification")
    notification_meta = notification_meta if isinstance(notification_meta, dict) else {}
    try:
        retry_count = max(0, int(notification_meta.get("retry_count") or 0))
    except (TypeError, ValueError):
        retry_count = 0
    if retry_count >= 1:
        log_event(
            logger,
            "whatsapp.retry_notification.skipped_max_retries",
            message_sid=sid,
            trigger=trigger,
        )
        return

    assignment_id = assignment.get("id")
    try:
        assignment_uuid = UUID(str(assignment_id))
    except (TypeError, ValueError):
        return

    notification_meta["retry_count"] = retry_count + 1
    notification_meta["retry_of_sid"] = sid
    notification_meta["retry_requested_at"] = datetime.now(timezone.utc).isoformat()
    if error_code:
        notification_meta["last_error_code"] = str(error_code)
    assignment_metadata = {**assignment_metadata, "notification": notification_meta}
    try:
        await repo.update_sales_assignment_notification(
            assignment_id=assignment_uuid,
            metadata=assignment_metadata,
        )
    except CRMRepositoryError as exc:
        log_event(
            logger,
            "whatsapp.retry_notification.metadata_update_failed",
            message_sid=sid,
            trigger=trigger,
            error=str(exc),
        )
        return

    conversation_id = str(assignment.get("conversacion_id") or "").strip()
    persona_id = str(assignment.get("contacto_id") or "").strip()
    opportunity_id = str(assignment.get("oportunidad_id") or "").strip() or None
    if not conversation_id or not persona_id:
        return

    try:
        contact = await storage.fetch_persona(persona_id)
    except StorageError as exc:
        log_event(
            logger,
            "whatsapp.retry_notification.contact_fetch_failed",
            message_sid=sid,
            trigger=trigger,
            error=str(exc),
        )
        return

    resumen = str(contact.get("necesidad_proposito") or "").strip() or None
    notes = str(contact.get("notes") or "").strip() or None
    email = _contact_email_value(contact)
    extra_reason = assignment_metadata.get("reason")
    extra: dict[str, Any] = dict(extra_reason) if isinstance(extra_reason, dict) else {}
    extra["retry_of_sid"] = sid
    extra["delivery_error_code"] = error_code

    channel_value = str(assignment.get("canal") or "").strip().lower() or "whatsapp"
    context = ToolRuntimeContext(
        conversation_id=conversation_id,
        persona_id=persona_id,
        channel=channel_value,
    )
    try:
        if channel_value == "webchat":
            from app.channels.webchat import notifications as webchat_notifications

            await webchat_notifications.notify_sales_rep(
                context=context,
                trigger=trigger,
                persona=contact,
                opportunity_id=opportunity_id,
                resumen=resumen,
                notes=notes,
                email=email,
                extra=extra,
                force_retry=True,
            )
        else:
            await whatsapp_tools._notify_sales_rep(
                context=context,
                trigger=trigger,
                persona=contact,
                opportunity_id=opportunity_id,
                resumen=resumen,
                notes=notes,
                email=email,
                extra=extra,
                force_retry=True,
            )
    except Exception as exc:  # pragma: no cover - best effort
        log_event(
            logger,
            "whatsapp.retry_notification.send_failed",
            message_sid=sid,
            trigger=trigger,
            channel=channel_value,
            error=str(exc),
        )
        return

    log_event(
        logger,
        "whatsapp.retry_notification.sent",
        message_sid=sid,
        trigger=trigger,
        channel=channel_value,
    )


async def _sync_envio_status_from_whatsapp(callback: schemas.WhatsAppStatusCallback) -> None:
    """Sincroniza el envío en prospección con el estatus de Twilio."""

    estado_envio = _map_status_to_envio_estado(callback.status)
    if not estado_envio:
        return
    try:
        repo = CRMRepository()
    except CRMRepositoryError as exc:
        log_event(logger, "whatsapp.status_envio_repo_error", error=str(exc))
        return
    try:
        envio = await repo.worker_get_envio_by_mensaje(mensaje_id=callback.message_sid)
    except CRMRepositoryError as exc:
        log_event(logger, "whatsapp.status_envio_fetch_failed", error=str(exc))
        return
    if not envio:
        return
    envio_id = envio.get("id")
    if not envio_id:
        return
    try:
        envio_uuid = UUID(str(envio_id))
    except (TypeError, ValueError):
        log_event(logger, "whatsapp.status_envio_invalid_id", envio_id=envio_id)
        return
    current_detalle = envio.get("detalle") if isinstance(envio.get("detalle"), dict) else {}
    merged_detalle = {
        **current_detalle,
        "status": callback.status,
        "timestamp": callback.timestamp,
        "error_code": callback.error_code,
    }
    payload = {
        "estado": estado_envio,
        "detalle": merged_detalle,
        "error": callback.error_code if estado_envio == "fallido" else None,
        "procesado_en": datetime.now(timezone.utc).isoformat(),
    }
    try:
        await repo.worker_complete_envio(envio_id=envio_uuid, payload=payload)
        metrics.increment("whatsapp", payload["estado"])
        batch_id_value = envio.get("batch_id")
        if batch_id_value:
            await progress_hub.publish(
                str(batch_id_value),
                {
                    "type": "envio",
                    "batch_id": batch_id_value,
                    "envio_id": str(envio_uuid),
                    "estado": payload["estado"],
                },
            )
        await repo.worker_insert_contact_logs(
            [
                {
                    "organizacion_id": (
                        str(envio.get("organizacion_id")) if envio.get("organizacion_id") else None
                    ),
                    "prospecto_id": (
                        str(envio.get("prospecto_id")) if envio.get("prospecto_id") else None
                    ),
                    "canal": "whatsapp",
                    "estado": estado_envio,
                    "detalle": {
                        "status": callback.status,
                        "timestamp": callback.timestamp,
                    },
                    "error": callback.error_code if estado_envio == "fallido" else None,
                    "batch_id": str(envio.get("batch_id")) if envio.get("batch_id") else None,
                    "envio_id": str(envio_uuid),
                }
            ]
        )
        await auto_promote_prospecto(
            prospecto_id=envio.get("prospecto_id"),
            canal="whatsapp",
            estado=estado_envio,
            repo=repo,
        )
        batch_state = None
        if estado_envio == "fallido" and batch_id_value:
            batch_state = await repo.worker_sync_batch_status(batch_id=UUID(str(batch_id_value)))
        if batch_state and batch_id_value:
            await progress_hub.publish(
                str(batch_id_value),
                {
                    "type": "batch",
                    "batch_id": batch_id_value,
                    "estado": batch_state,
                },
            )
    except CRMRepositoryError as exc:
        log_event(
            logger,
            "whatsapp.status_envio_update_failed",
            error=str(exc),
            message_sid=callback.message_sid,
        )


async def _maybe_update_persona_location(
    persona_id: str,
    persona: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    """Enriquece la persona con la ubicación inferida a partir de su teléfono/LADA."""
    persona_data = persona
    if persona_data is None:
        try:
            persona_data = await storage.fetch_persona(persona_id)
        except StorageError as exc:
            logger.warning(
                "whatsapp.fetch_persona_failed",
                extra={"persona_id": persona_id, "error": str(exc)},
            )
            return None

    persona_datos = _persona_datos(persona_data)
    ubicacion = dict(persona_datos.get("ubicacion") or {})
    lada_exists = ubicacion.get("lada")
    estado_exists = ubicacion.get("cve_ent")
    cvegeo_exists = ubicacion.get("cvegeo")

    if lada_exists and estado_exists and cvegeo_exists:
        return persona_data

    try:
        identities = await storage.fetch_persona_identities(persona_id)
    except StorageError as exc:
        logger.warning(
            "whatsapp.fetch_persona_identities_failed",
            extra={"persona_id": persona_id, "error": str(exc)},
        )
        identities = []

    channels = []
    origen = persona_data.get("origen")
    if isinstance(origen, str) and origen:
        channels.append(origen)
    else:
        channels.append("whatsapp")

    location = leads_geo.infer_contact_location(
        contacto_id=persona_id,
        data=persona_data,
        channels=channels,
        identities=identities,
    )

    updated = False
    ubicacion.setdefault("pais", "México")
    ubicacion.setdefault("country_code", "MX")

    if location.lada and ubicacion.get("lada") != location.lada:
        ubicacion["lada"] = location.lada
        updated = True
    if location.estado_clave and ubicacion.get("cve_ent") != location.estado_clave:
        ubicacion["cve_ent"] = location.estado_clave
        updated = True
    if location.estado_nombre and ubicacion.get("nom_ent") != location.estado_nombre:
        ubicacion["nom_ent"] = location.estado_nombre
        updated = True
    if location.municipio_clave and ubicacion.get("cve_mun") != location.municipio_clave:
        ubicacion["cve_mun"] = location.municipio_clave
        updated = True
    if location.municipio_nombre and ubicacion.get("nom_mun") != location.municipio_nombre:
        ubicacion["nom_mun"] = location.municipio_nombre
        updated = True
    if location.municipio_cvegeo and ubicacion.get("cvegeo") != location.municipio_cvegeo:
        ubicacion["cvegeo"] = location.municipio_cvegeo
        updated = True

    if not updated:
        return persona_data

    persona_datos["ubicacion"] = ubicacion
    try:
        await storage.update_persona(persona_id, {"persona_datos": persona_datos})
    except StorageError as exc:
        logger.warning(
            "whatsapp.update_persona_location_failed",
            extra={"persona_id": persona_id, "error": str(exc)},
        )
    else:
        persona_data["persona_datos"] = persona_datos
        persona_data["contacto_datos"] = dict(persona_datos)

    return persona_data

async def _generate_assistant_reply(
    *,
    message: schemas.WhatsAppIncomingMessage,
    conversation_id: str,
    persona_id: str,
    openai_conversation_id: str | None,
    previous_response_id: str | None,
    catalog_context: str | None,
    booking_context: str | None,
    whatsapp_settings: tenant_runtime.WhatsappRuntimeSettings,
    organizacion_id: UUID | None,
    catalog_inmobiliario_enabled: bool = True,
    catalog_no_inmobiliario_enabled: bool = True,
    prospeccion_mode: bool = False,
    origin_type: str | None = None,
    inbound_message_id: str | None = None,
    attachment_content_items: list[dict[str, Any]] | None = None,
) -> AssistantReply:
    debug_timings: dict[str, float] = {}
    started = time.perf_counter()
    assistant = _build_assistant_from_runtime(whatsapp_settings, prospeccion_mode=prospeccion_mode)
    debug_timings["assistant_build_ms"] = round((time.perf_counter() - started) * 1000, 2)
    log_event(
        logger,
        "whatsapp.assistant_routing",
        prospeccion_mode=prospeccion_mode,
        origin_type=origin_type,
        prompt_id=assistant.prompt_id,
        prompt_version=assistant.prompt_version,
        assistant_id=assistant.assistant_id,
    )
    client = openai_service.get_assistant_client(
        api_key=whatsapp_settings.voice_api_key,
        project_id=whatsapp_settings.project_id,
    )
    metadata_payload = _build_whatsapp_openai_metadata_payload(
        conversation_id=conversation_id,
        persona_id=persona_id,
        message_sid=message.message_sid,
        inbound_message_id=inbound_message_id,
        prospeccion_mode=prospeccion_mode,
        origin_type=origin_type,
    )
    assistant_spec = None
    if not assistant.is_prompt:
        if not assistant.assistant_id:
            raise RuntimeError("WHATSAPP_ASSISTANT_ID is not configured")
        assistant_spec_started = time.perf_counter()
        assistant_spec = await resolve_assistant_spec(client, assistant.assistant_id)
        debug_timings["assistant_spec_ms"] = round((time.perf_counter() - assistant_spec_started) * 1000, 2)
    filtered_assistant_tools = (
        filter_assistant_tools(
            assistant_spec.tools,
            catalog_inmobiliario_enabled=catalog_inmobiliario_enabled,
            catalog_no_inmobiliario_enabled=catalog_no_inmobiliario_enabled,
        )
        if assistant_spec
        else []
    )
    context_payload: dict[str, Any] | None = None
    try:
        context_fetch_started = time.perf_counter()
        context_payload = await storage.fetch_persona_context(
            conversation_id=conversation_id,
            persona_id=persona_id,
        )
        debug_timings["fetch_persona_context_ms"] = round((time.perf_counter() - context_fetch_started) * 1000, 2)
    except StorageError as exc:  # pragma: no cover - fallbacks informativos
        logger.warning(
            "whatsapp.fetch_persona_context_failed",
            extra={
                "conversation_id": conversation_id,
                "persona_id": persona_id,
                "error": str(exc),
            },
        )

    summary_record: dict[str, Any] | None = None
    summary_text: str | None = None
    summary_created_en: str | None = None
    try:
        summary_started = time.perf_counter()
        summary_record = await conversation_summary.ensure_conversation_summary(
            conversation_id=conversation_id,
            persona_id=persona_id,
            context_data=context_payload,
            organizacion_id=organizacion_id,
            generate_if_missing=False,
        )
        debug_timings["summary_initial_ms"] = round((time.perf_counter() - summary_started) * 1000, 2)
        if summary_record:
            candidate = summary_record.get("resumen")
            if isinstance(candidate, str) and candidate.strip():
                summary_text = candidate.strip()
            summary_created_en = summary_record.get("creado_en")
            metadata = summary_record.get("metadatos")
            if isinstance(metadata, dict) and metadata:
                metadata = {k: v for k, v in metadata.items() if k != "type"}
            else:
                metadata = {}
            summary_record["metadatos"] = metadata
    except Exception as exc:  # pragma: no cover
        logger.warning(
            "whatsapp.conversation_summary_failed",
            extra={"conversation_id": conversation_id, "error": str(exc)},
        )

    if booking_context:
        context_payload = context_payload or {}
        context_payload["booking_context"] = booking_context
    if origin_type:
        context_payload = context_payload or {}
        context_payload["origin_type"] = origin_type

    initial_input = _build_openai_input(
        message,
        context_data=context_payload,
        summary_text=summary_text,
        summary_created_en=summary_created_en,
        catalog_context=catalog_context,
        attachment_content_items=attachment_content_items,
    )
    profiling_enabled_for_channel = True
    if organizacion_id:
        try:
            profiling_started = time.perf_counter()
            profiling_enabled_for_channel = await tenant_runtime.is_profiling_enabled(
                organizacion_id=organizacion_id,
                channel="whatsapp",
            )
            debug_timings["profiling_toggle_ms"] = round((time.perf_counter() - profiling_started) * 1000, 2)
        except Exception as exc:  # pragma: no cover
            logger.warning(
                "whatsapp.profiling_toggle_lookup_failed",
                extra={
                    "conversation_id": conversation_id,
                    "organizacion_id": str(organizacion_id),
                    "error": str(exc),
                },
            )
    initial_input.insert(
        0,
        {
            "role": "developer",
            "content": [
                {
                    "type": "input_text",
                    "text": (
                        "Fecha y hora del servidor: "
                        + get_current_time_reference()
                        + " Si el usuario pregunta por la fecha actual, responde con esta información."
                    ),
                }
            ],
        },
    )
    initial_input.insert(
        1,
        {
            "role": "developer",
            "content": [
                {
                    "type": "input_text",
                    "text": (
                        "Contrato operativo para close_lead: cuando captures perfilamiento, incluye "
                        "profiling_statuses y profiling_reprompt_counts por campo. "
                        "profiling_statuses debe usar solo: answered, unknown, refused, skipped_max_retries. "
                        "Si un campo no se obtuvo tras la repregunta máxima, marca skipped_max_retries y continua. "
                        "No forces al usuario con repreguntas adicionales."
                        if profiling_enabled_for_channel
                        else "Perfilamiento IA desactivado para este tenant/canal. "
                        "No hagas preguntas de perfilamiento o scoring. "
                        "No envíes campos de scoring en close_lead "
                        "(financing_type, credit_preapproved, budget_range, purchase_timeline, "
                        "decision_authority, visited_properties, profiling_statuses, profiling_reprompt_counts). "
                        "Flujo permitido: captura de datos básicos, close_lead simple, "
                        "y gestión de agenda/email si el prospecto lo pide."
                    ),
                }
            ],
        },
    )
    initial_input.insert(
        2,
        {
            "role": "developer",
            "content": [
                {
                    "type": "input_text",
                    "text": (
                        "Estilo WhatsApp (regla estricta): responde breve. "
                        "Por defecto usa 1–3 frases (máx. ~300 caracteres) y termina con 1 pregunta. "
                        "No des listas largas ni autopromoción; ofrece ampliar solo si el usuario pide detalles."
                    ),
                }
            ],
        },
    )
    location_prompt = (
        "Regla de ubicación comercial: la ubicación del Contexto CRM (incluida LADA) "
        "es solo referencia técnica y no define la zona de búsqueda del prospecto. "
        "Nunca preguntes si busca en la zona inferida por su teléfono. "
    )
    inventory_prompt = (
        "Si el prospecto menciona una zona/fraccionamiento sin coincidencias claras, "
        "ejecuta list_catalog_fraccionamientos para obtener inventario real y responde "
        "con zonas/fraccionamientos disponibles antes de hacer una sola pregunta de avance."
        if catalog_inmobiliario_enabled
        else "Catálogo inmobiliario desactivado para este tenant/canal. "
        "No uses recursos de catálogo inmobiliario ni infieras inventario desde ellos. "
        "Si falta contexto, haz una sola pregunta de aclaración usando solo la información "
        "permitida por el tenant."
    )
    product_catalog_prompt = (
        "Si el prospecto pregunta por productos, servicios o paquetes, usa fetch_catalog_item_details "
        "para buscar el catálogo comercial general antes de responder."
        if catalog_no_inmobiliario_enabled
        else "Catálogo de productos y servicios desactivado para este tenant/canal. "
        "No uses recursos del catálogo comercial general."
    )
    initial_input.insert(
        3,
        {
            "role": "developer",
            "content": [{"type": "input_text", "text": location_prompt + inventory_prompt}],
        },
    )
    initial_input.insert(
        4,
        {
            "role": "developer",
            "content": [{"type": "input_text", "text": product_catalog_prompt}],
        },
    )
    initial_input.insert(
        5,
        {
            "role": "developer",
            "content": [
                {
                    "type": "input_text",
                    "text": (
                        "Nombre real del lead: en `set_full_name` usa solo el nombre que el usuario "
                        "escribió explícitamente en el chat. Nunca uses el nombre del perfil de WhatsApp, "
                        "alias genéricos ni placeholders como 'Visitante WhatsApp'."
                    ),
                }
            ],
        },
    )
    location_href = getattr(whatsapp_settings, "location_href", None)
    if location_href:
        initial_input.insert(
            6,
            {
                "role": "developer",
                "content": [
                    {
                        "type": "input_text",
                        "text": "\n".join(build_location_context_lines(location_href)),
                    }
                ],
            },
        )
    if booking_context:
        initial_input.insert(
            3,
            {
                "role": "developer",
                "content": [
                    {
                        "type": "input_text",
                        "text": booking_context,
                    }
                ],
            },
        )
    if prospeccion_mode:
        initial_input.insert(
            3,
            {
                "role": "developer",
                "content": [
                    {
                        "type": "input_text",
                        "text": (
                            "Regla de agenda en prospección: nunca digas que no puedes agendar. "
                            "Si el prospecto confirma fecha/hora o acepta cita, DEBES usar tools "
                            "(list_demo_slots y/o schedule_demo) antes de responder. "
                            "Si falta algún dato requerido, pide solo ese dato faltante en una pregunta corta. "
                            "Orden obligatorio de captura antes de agendar: nombre completo, luego correo, luego empresa. "
                            "Si el usuario ya dio uno o varios datos en su(s) último(s) mensaje(s), extráelos y guárdalos "
                            "con tools de captura antes de preguntar; no los vuelvas a pedir."
                        ),
                    }
                ],
            },
        )
        initial_input.insert(
            4,
            {
                "role": "developer",
                "content": [
                    {
                        "type": "input_text",
                        "text": (
                            "Si ya existe cita confirmada en el contexto, no reinicies calificación ni agenda. "
                            "No vuelvas a pedir nombre/correo/empresa. "
                            "Solo confirma y responde dudas puntuales; "
                            "si el usuario quiere cambiar/cancelar, usa reschedule_demo o cancel_demo."
                        ),
                    }
                ],
            },
        )
        initial_input.insert(
            5,
            {
                "role": "developer",
                "content": [
                    {
                        "type": "input_text",
                        "text": (
                            "Evita preguntas de reconfirmación redundantes. "
                            "Solo confirma un dato cuando sea ambiguo o potencialmente inválido "
                            "(por ejemplo, correo con formato dudoso). "
                            "Si nombre/correo/empresa ya están claros en contexto, continúa sin pedirlos otra vez."
                        ),
                    }
                ],
            },
        )
        initial_input.insert(
            6,
            {
                "role": "developer",
                "content": [
                    {
                        "type": "input_text",
                        "text": (
                            "origin_type de esta conversación: "
                            f"{str(origin_type or 'prospeccion').strip().lower()}. "
                            "Si es publicidad_whatsapp, llegó por frase atribuida. "
                            "Si es prospeccion, llegó por campaña/plantilla."
                        ),
                    }
                ],
            },
        )
    wants_detail = _wants_detailed_reply(message.body)
    # Este presupuesto aplica al loop de tools; debe ser holgado para evitar
    # truncar JSON de function calls.
    max_output_tokens = 1200 if wants_detail else 900
    request_kwargs: dict[str, Any] = {
        "input": initial_input,
        "store": True,
        "max_output_tokens": max_output_tokens,
        "temperature": 0.4,
        "metadata": metadata_payload,
        "tool_choice": "auto",
    }

    prompt_variables: dict[str, Any] = {"conversacion_id": conversation_id}
    if location_href:
        prompt_variables["location_href"] = location_href

    def _build_request_template(*, include_tools: bool = True) -> dict[str, Any]:
        if assistant.is_prompt:
            return {
                "prompt": build_prompt_payload(assistant, prompt_variables),
                "text": {"format": {"type": "text"}},
            }
        if not assistant_spec:
            raise RuntimeError("No se pudo resolver la configuración del asistente")
        payload: dict[str, Any] = {"model": assistant_spec.model}
        if assistant_spec.instructions:
            payload["instructions"] = assistant_spec.instructions
        if include_tools and filtered_assistant_tools:
            payload["tools"] = filtered_assistant_tools
        return payload

    request_kwargs.update(_build_request_template(include_tools=True))
    if assistant.is_prompt:
        request_kwargs.pop("temperature", None)

    context_obj = ToolRuntimeContext(
        conversation_id=conversation_id,
        persona_id=persona_id,
        session_id=f"whatsapp:{conversation_id}",
        channel="whatsapp",
        organizacion_id=str(organizacion_id) if organizacion_id else None,
        feature="sales_chat",
        catalog_inmobiliario_enabled=catalog_inmobiliario_enabled,
        catalog_no_inmobiliario_enabled=catalog_no_inmobiliario_enabled,
    )

    async def _run_assistant_generation(current_previous_response_id: str | None) -> tuple[Any, float]:
        run_request_kwargs = dict(request_kwargs)
        run_request_kwargs.update(_build_request_template(include_tools=True))
        if assistant.is_prompt:
            run_request_kwargs.pop("temperature", None)
        if openai_conversation_id:
            run_request_kwargs["conversation"] = openai_conversation_id
        elif current_previous_response_id:
            run_request_kwargs["previous_response_id"] = current_previous_response_id
        else:
            run_request_kwargs.pop("previous_response_id", None)
        tool_loop_started = time.perf_counter()
        result = await run_tool_loop(
            client=client,
            assistant=assistant,
            assistant_spec=assistant_spec,
            context=context_obj,
            initial_request=run_request_kwargs,
            request_template=lambda: _build_request_template(include_tools=True),
            execute_tool=whatsapp_tools.execute_tool,
            openai_conversation_id=openai_conversation_id,
            previous_response_id=current_previous_response_id,
            api_key=whatsapp_settings.voice_api_key,
            log=logger,
        )
        return result, round((time.perf_counter() - tool_loop_started) * 1000, 2)

    try:
        result, tool_loop_ms = await _run_assistant_generation(previous_response_id)
        debug_timings["tool_loop_ms"] = tool_loop_ms
        tool_runtime_debug = result.side_effects.get("tool_runtime_debug")
        if isinstance(tool_runtime_debug, dict) and tool_runtime_debug:
            debug_timings["tool_runtime_debug"] = tool_runtime_debug
    except Exception as exc:
        if previous_response_id and _is_previous_response_not_found_error(exc):
            log_event(
                logger,
                "whatsapp.previous_response_missing_retry",
                conversation_id=conversation_id,
                previous_response_id=_trim_text(previous_response_id),
            )
            previous_response_id = None
            result, tool_loop_retry_ms = await _run_assistant_generation(None)
            debug_timings["tool_loop_retry_ms"] = tool_loop_retry_ms
            tool_runtime_debug = result.side_effects.get("tool_runtime_debug")
            if isinstance(tool_runtime_debug, dict) and tool_runtime_debug:
                debug_timings["tool_runtime_debug"] = tool_runtime_debug
        elif assistant.is_prompt and prompt_variables and _is_unknown_prompt_variable_error(exc):
            log_event(
                logger,
                "whatsapp.prompt_variables_retry_without_variables",
                conversation_id=conversation_id,
                prompt_id=assistant.prompt_id,
                prompt_version=assistant.prompt_version,
            )
            prompt_variables = {"conversacion_id": prompt_variables.get("conversacion_id", conversation_id)}
            result, tool_loop_retry_ms = await _run_assistant_generation(previous_response_id)
            debug_timings["tool_loop_retry_ms"] = tool_loop_retry_ms
            tool_runtime_debug = result.side_effects.get("tool_runtime_debug")
            if isinstance(tool_runtime_debug, dict) and tool_runtime_debug:
                debug_timings["tool_runtime_debug"] = tool_runtime_debug
        else:
            raise

    reply_text = _extract_text_from_response(result.response)
    final_text = _compact_whatsapp_reply(reply_text, _DEFAULT_WHATSAPP_MAX_CHARS)
    final_response_id = result.response_id
    final_conversation_id = result.conversation_id
    welcome_document_sent_by_tool = bool(result.side_effects.get("welcome_document_sent")) or (
        "send_information_package" in result.tools_called
    )

    quality_eval_started = time.perf_counter()
    quality_ok, quality_reason = evaluate_reply_quality(final_text)
    debug_timings["quality_eval_ms"] = round((time.perf_counter() - quality_eval_started) * 1000, 2)
    if not quality_ok:
        logger.warning(
            "whatsapp.reply_quality_low",
            extra={"conversation_id": conversation_id, "reason": quality_reason},
        )
        if quality_reason == "empty":
            final_text = None
            log_event(
                logger,
                "whatsapp.empty_reply_payload",
                conversation_id=conversation_id,
                inbound_message_id=inbound_message_id,
                assistant_response_id=final_response_id,
                response_summary=_summarize_openai_response_payload(result.response),
            )
            logger.info(
                "whatsapp.reply_quality_skip_retry",
                extra={"conversation_id": conversation_id, "reason": quality_reason},
            )
        else:
            guard_retry_kwargs: dict[str, Any] = {
                "input": [
                    {
                        "role": "developer",
                        "content": [
                            {
                                "type": "input_text",
                                "text": (
                                    "Regenera SOLO un mensaje final de WhatsApp completo y autocontenido. "
                                    "1-3 frases, máximo 500 caracteres. "
                                    "No termines con puntos suspensivos, comas ni conectores sueltos."
                                ),
                            }
                        ],
                    }
                ],
                "store": True,
                "max_output_tokens": 180,
                "temperature": 0.2,
                "metadata": metadata_payload,
                "tool_choice": "none",
            }
            guard_retry_kwargs.update(_build_request_template(include_tools=False))
            if assistant.is_prompt:
                guard_retry_kwargs.pop("temperature", None)
            if final_conversation_id:
                guard_retry_kwargs["conversation"] = final_conversation_id
            elif final_response_id:
                guard_retry_kwargs["previous_response_id"] = final_response_id
            try:
                quality_retry_started = time.perf_counter()
                retry_response = await client.responses.create(**guard_retry_kwargs)
                debug_timings["quality_retry_ms"] = round((time.perf_counter() - quality_retry_started) * 1000, 2)
                retry_payload = retry_response.model_dump()
                await openai_usage_ledger.record_response_usage(
                    organizacion_id=organizacion_id,
                    channel="whatsapp",
                    feature="sales_chat",
                    assistant=assistant,
                    response_payload=retry_payload,
                    request_purpose="quality_retry",
                    latency_ms=int(round(debug_timings["quality_retry_ms"])),
                    api_key=whatsapp_settings.voice_api_key,
                    request_metadata={"conversation_id": conversation_id},
                    conversation_id=conversation_id,
                    persona_id=persona_id,
                    quality_retry_used=True,
                    project_id=assistant.project_id,
                )
                retry_text = _extract_text_from_response(retry_payload)
                retry_ok, retry_reason = evaluate_reply_quality(retry_text)
                if retry_ok:
                    final_text = retry_text
                    final_response_id = retry_payload.get("id") or final_response_id
                    final_conversation_id = (
                        (retry_payload.get("conversation") or {}).get("id")
                        or final_conversation_id
                    )
                    logger.info(
                        "whatsapp.reply_quality_recovered",
                        extra={"conversation_id": conversation_id, "previous_reason": quality_reason},
                    )
                else:
                    final_text = None
                    logger.warning(
                        "whatsapp.reply_quality_retry_failed",
                        extra={
                            "conversation_id": conversation_id,
                            "previous_reason": quality_reason,
                            "retry_reason": retry_reason,
                        },
                    )
            except Exception as exc:  # pragma: no cover
                final_text = None
                logger.warning(
                    "whatsapp.reply_quality_retry_exception",
                    extra={
                        "conversation_id": conversation_id,
                        "previous_reason": quality_reason,
                        "error": str(exc),
                    },
                )

    if conversation_id:
        asyncio.create_task(
            _refresh_conversation_summary_best_effort(
                conversation_id=conversation_id,
                persona_id=persona_id,
                organizacion_id=organizacion_id,
                context_data=context_payload,
            )
        )
        debug_timings["summary_refresh_scheduled"] = 1.0

    return AssistantReply(
        text=final_text.strip() if final_text else None,
        openai_conversation_id=final_conversation_id,
        response_id=final_response_id,
        tools_called=result.tools_called,
        debug_timings=debug_timings,
    )


async def _send_twilio_whatsapp_reply(
    *,
    to_number: str,
    body: str | None = None,
    content_sid: str | None = None,
    content_variables: dict[str, str] | None = None,
    attachments: list[dict[str, Any]] | None = None,
    organizacion_id: UUID | None = None,
) -> TwilioSendResult:
    """Envía la respuesta al contacto utilizando la API de Twilio."""

    runtime = await tenant_runtime.get_twilio_runtime_settings(organizacion_id=organizacion_id)
    if (
        not runtime.phone_number
        or not runtime.account_sid
        or not runtime.auth_token
    ):
        logger.warning("whatsapp.twilio_not_configured")
        return TwilioSendResult(sid=None, status="skipped", error="twilio_not_configured")

    if not body and not content_sid and not attachments:
        logger.warning("whatsapp.empty_payload")
        return TwilioSendResult(sid=None, status="skipped", error="empty_payload")

    normalized_to = to_number or ""
    if normalized_to and not normalized_to.lower().startswith("whatsapp:"):
        normalized_to = f"whatsapp:{normalized_to}"
    normalized_from = runtime.phone_number
    if normalized_from and not normalized_from.lower().startswith("whatsapp:"):
        normalized_from = f"whatsapp:{normalized_from}"

    client = twilio_service.get_twilio_client_for_credentials(runtime.account_sid, runtime.auth_token)
    message_kwargs: dict[str, Any] = {
        "to": normalized_to,
        "from_": normalized_from,
    }
    if content_sid:
        message_kwargs["content_sid"] = content_sid
        if content_variables:
            message_kwargs["content_variables"] = json.dumps(
                content_variables,
                ensure_ascii=False,
            )
    else:
        message_kwargs["body"] = body or ""
        if attachments:
            first_attachment = next((item for item in attachments if isinstance(item, dict) and item.get("url")), None)
            if first_attachment and first_attachment.get("url"):
                message_kwargs["media_url"] = [str(first_attachment["url"])]

    try:
        message = await asyncio.to_thread(
            client.messages.create,
            **message_kwargs,
        )
    except Exception as exc:  # pragma: no cover - errores propios del SDK
        logger.exception("whatsapp.twilio_send_failed", extra={"error": str(exc)})
        return TwilioSendResult(sid=None, status="failed", error=str(exc))

    status = getattr(message, "status", None)
    return TwilioSendResult(
        sid=getattr(message, "sid", None),
        status=status,
        error=None,
        from_number=normalized_from,
        provider="twilio",
    )


async def _send_meta_whatsapp_reply(
    *,
    to_number: str,
    body: str | None = None,
    content_sid: str | None = None,
    content_variables: dict[str, str] | None = None,
    template_name: str | None = None,
    template_language: str | None = None,
    attachments: list[dict[str, Any]] | None = None,
    organizacion_id: UUID | None = None,
) -> TwilioSendResult:
    """Envía la respuesta al contacto utilizando WhatsApp Cloud API."""

    runtime = await tenant_runtime.get_whatsapp_runtime_settings(
        organizacion_id=organizacion_id,
    )
    if not runtime.meta_phone_number_id or not runtime.meta_page_access_token:
        logger.warning("whatsapp.meta_not_configured")
        return TwilioSendResult(sid=None, status="skipped", error="meta_not_configured", provider="meta")

    normalized_template_name = _normalize_meta_template_name(template_name)
    normalized_template_language = _normalize_meta_template_language(template_language)
    if normalized_template_name or normalized_template_language:
        if not normalized_template_name or not normalized_template_language:
            logger.warning(
                "whatsapp.meta_template_incomplete",
                extra={
                    "template_name": template_name,
                    "template_language": template_language,
                },
            )
            return TwilioSendResult(
                sid=None,
                status="skipped",
                error="invalid_template",
                provider="meta",
            )
    else:
        if content_sid:
            logger.warning("whatsapp.meta_template_not_supported", extra={"content_sid": content_sid})
    if not body and not attachments and not normalized_template_name:
        logger.warning("whatsapp.empty_payload")
        return TwilioSendResult(sid=None, status="skipped", error="empty_payload", provider="meta")

    normalized_to = _normalize_meta_recipient_number(to_number)
    if not normalized_to:
        return TwilioSendResult(sid=None, status="skipped", error="invalid_recipient", provider="meta")

    graph_version = runtime.meta_graph_api_version or "v21.0"
    messages_url = f"https://graph.facebook.com/{graph_version}/{runtime.meta_phone_number_id}/messages"
    if normalized_template_name and normalized_template_language:
        payload = {
            "messaging_product": "whatsapp",
            "to": normalized_to,
            **_build_meta_template_payload(
                template_name=normalized_template_name,
                template_language=normalized_template_language,
                content_variables=content_variables,
            ),
        }
    else:
        attachment = next((item for item in attachments or [] if isinstance(item, dict) and item.get("url")), None)
        if attachment:
            mime = str(attachment.get("mime") or "").lower()
            attachment_url = str(attachment["url"])
            name = str(attachment.get("name") or "").strip()
            if mime.startswith("image/"):
                payload = {
                    "messaging_product": "whatsapp",
                    "to": normalized_to,
                    "type": "image",
                    "image": {"link": attachment_url, **({"caption": body} if body else {})},
                }
            elif mime.startswith("audio/"):
                payload = {
                    "messaging_product": "whatsapp",
                    "to": normalized_to,
                    "type": "audio",
                    "audio": {"link": attachment_url},
                }
            elif mime.startswith("video/"):
                payload = {
                    "messaging_product": "whatsapp",
                    "to": normalized_to,
                    "type": "video",
                    "video": {"link": attachment_url, **({"caption": body} if body else {})},
                }
            else:
                media_id = await _upload_meta_whatsapp_document(
                    runtime=runtime,
                    source_url=attachment_url,
                    filename=name or "documento.pdf",
                    mime=mime or "application/pdf",
                )
                if media_id:
                    payload = {
                        "messaging_product": "whatsapp",
                        "to": normalized_to,
                        "type": "document",
                        "document": {
                            "id": media_id,
                            **({"caption": body} if body else {}),
                            **({"filename": name} if name else {}),
                        },
                    }
                else:
                    payload = {
                        "messaging_product": "whatsapp",
                        "to": normalized_to,
                        "type": "document",
                        "document": {
                            "link": attachment_url,
                            **({"caption": body} if body else {}),
                            **({"filename": name} if name else {}),
                        },
                    }
        else:
            payload = {
                "messaging_product": "whatsapp",
                "to": normalized_to,
                "type": "text",
                "text": {"body": body or ""},
            }
    headers = {
        "Authorization": f"Bearer {runtime.meta_page_access_token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }

    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.post(messages_url, json=payload, headers=headers)
    except Exception as exc:  # pragma: no cover - red/servicio externo
        logger.exception("whatsapp.meta_send_failed", extra={"error": str(exc)})
        return TwilioSendResult(sid=None, status="failed", error=str(exc), provider="meta")

    if response.status_code >= 400:
        logger.warning(
            "whatsapp.meta_reply_failed",
            extra={
                "status_code": response.status_code,
                "body": response.text,
                "recipient": normalized_to,
            },
        )
        return TwilioSendResult(
            sid=None,
            status="failed",
            error=f"http_{response.status_code}",
            provider="meta",
        )

    response_payload: dict[str, Any] = {}
    try:
        response_payload = response.json()
    except ValueError:
        response_payload = {}
    message_id = None
    messages = response_payload.get("messages") if isinstance(response_payload, dict) else None
    if isinstance(messages, list) and messages and isinstance(messages[0], dict):
        message_id = _trim_text(messages[0].get("id"))
    if not message_id and isinstance(response_payload, dict):
        message_id = _trim_text(response_payload.get("message_id") or response_payload.get("id"))

    return TwilioSendResult(
        sid=message_id,
        status="sent",
        error=None,
        from_number=runtime.meta_phone_number_id,
        provider="meta",
    )


async def _upload_meta_whatsapp_document(
    *,
    runtime: Any,
    source_url: str,
    filename: str,
    mime: str,
) -> str | None:
    """Sube un PDF a Meta para enviarlo luego como documento por ID."""

    if not runtime.meta_phone_number_id or not runtime.meta_page_access_token:
        return None
    graph_version = runtime.meta_graph_api_version or "v21.0"
    upload_url = f"https://graph.facebook.com/{graph_version}/{runtime.meta_phone_number_id}/media"
    try:
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            source_response = await client.get(source_url)
            if source_response.status_code >= 400:
                return None
            upload_response = await client.post(
                upload_url,
                data={
                    "messaging_product": "whatsapp",
                    "type": mime or "application/pdf",
                },
                files={
                    "file": (filename, source_response.content, mime or "application/pdf"),
                },
                headers={
                    "Authorization": f"Bearer {runtime.meta_page_access_token}",
                    "Accept": "application/json",
                },
            )
    except Exception:
        return None

    if upload_response.status_code >= 400:
        logger.warning(
            "whatsapp.meta_media_upload_failed",
            extra={"status_code": upload_response.status_code, "body": upload_response.text},
        )
        return None

    try:
        data = upload_response.json()
    except ValueError:
        return None
    media_id = data.get("id") if isinstance(data, dict) else None
    return str(media_id).strip() if media_id else None


async def _send_whatsapp_reply(
    *,
    to_number: str,
    body: str | None = None,
    content_sid: str | None = None,
    content_variables: dict[str, str] | None = None,
    template_name: str | None = None,
    template_language: str | None = None,
    attachments: list[dict[str, Any]] | None = None,
    organizacion_id: UUID | None = None,
) -> TwilioSendResult:
    runtime = await tenant_runtime.get_whatsapp_runtime_settings(
        organizacion_id=organizacion_id,
    )
    provider = _normalize_whatsapp_provider(runtime.provider)
    if provider == "meta":
        return await _send_meta_whatsapp_reply(
            to_number=to_number,
            body=body,
            content_sid=content_sid,
            content_variables=content_variables,
            template_name=template_name,
            template_language=template_language,
            attachments=attachments,
            organizacion_id=organizacion_id,
        )
    return await _send_twilio_whatsapp_reply(
        to_number=to_number,
        body=body,
        content_sid=content_sid,
        content_variables=content_variables,
        attachments=attachments,
        organizacion_id=organizacion_id,
    )


async def _send_whatsapp_typing_indicator(
    *,
    incoming_message_sid: str | None,
    organizacion_id: UUID | None,
) -> bool:
    """Dispara el indicador de escritura para WhatsApp vía Twilio.

    Es best-effort: nunca debe romper el flujo de respuesta principal.
    """

    message_sid = str(incoming_message_sid or "").strip()
    if not message_sid:
        return False
    if organizacion_id is None:
        return False

    runtime = await tenant_runtime.get_whatsapp_runtime_settings(
        organizacion_id=organizacion_id,
    )
    provider = _normalize_whatsapp_provider(runtime.provider)
    if provider == "meta":
        return False
    runtime_twilio = await tenant_runtime.get_twilio_runtime_settings(organizacion_id=organizacion_id)
    if not runtime_twilio.account_sid or not runtime_twilio.auth_token:
        return False

    try:
        async with httpx.AsyncClient(timeout=_WHATSAPP_TYPING_TIMEOUT_SECONDS) as client:
            response = await client.post(
                _WHATSAPP_TYPING_INDICATOR_URL,
                data={"messageId": message_sid, "channel": "whatsapp"},
                auth=(runtime_twilio.account_sid, runtime_twilio.auth_token),
                headers={"Accept": "application/json"},
            )
        if 200 <= response.status_code < 300:
            log_event(logger, "whatsapp.typing_indicator_sent", message_sid=message_sid)
            return True
        logger.info(
            "whatsapp.typing_indicator_not_sent",
            extra={
                "message_sid": message_sid,
                "status_code": response.status_code,
                "response_preview": response.text[:300],
            },
        )
        return False
    except Exception as exc:  # pragma: no cover - red/servicio externo
        logger.info(
            "whatsapp.typing_indicator_failed",
            extra={"message_sid": message_sid, "error": str(exc)},
        )
        return False


async def _send_whatsapp_read_indicator(
    *,
    incoming_message_sid: str | None,
    organizacion_id: UUID | None,
) -> bool:
    """Marca como leído el mensaje entrante vía Twilio (best-effort)."""

    message_sid = str(incoming_message_sid or "").strip()
    if not message_sid:
        return False
    if organizacion_id is None:
        return False

    runtime = await tenant_runtime.get_whatsapp_runtime_settings(
        organizacion_id=organizacion_id,
        force_refresh=True,
    )
    provider = _normalize_whatsapp_provider(runtime.provider)
    if provider == "meta":
        if not runtime.meta_phone_number_id or not runtime.meta_page_access_token:
            return False
        graph_version = runtime.meta_graph_api_version or "v21.0"
        url = f"https://graph.facebook.com/{graph_version}/{runtime.meta_phone_number_id}/messages"
        payload = {
            "messaging_product": "whatsapp",
            "status": "read",
            "message_id": message_sid,
        }
        headers = {
            "Authorization": f"Bearer {runtime.meta_page_access_token}",
            "Content-Type": "application/json",
            "Accept": "application/json",
        }
        try:
            async with httpx.AsyncClient(timeout=_WHATSAPP_READ_TIMEOUT_SECONDS) as client:
                response = await client.post(url, json=payload, headers=headers)
            if 200 <= response.status_code < 300:
                log_event(logger, "whatsapp.read_indicator_sent", message_sid=message_sid)
                return True
            logger.info(
                "whatsapp.read_indicator_not_sent",
                extra={
                    "message_sid": message_sid,
                    "status_code": response.status_code,
                    "response_preview": response.text[:300],
                },
            )
            return False
        except Exception as exc:  # pragma: no cover - red/servicio externo
            logger.info(
                "whatsapp.read_indicator_failed",
                extra={"message_sid": message_sid, "error": str(exc)},
            )
            return False

    runtime_twilio = await tenant_runtime.get_twilio_runtime_settings(organizacion_id=organizacion_id)
    if not runtime_twilio.account_sid or not runtime_twilio.auth_token:
        return False

    try:
        async with httpx.AsyncClient(timeout=_WHATSAPP_READ_TIMEOUT_SECONDS) as client:
            response = await client.post(
                _WHATSAPP_READ_INDICATOR_URL,
                data={"messageId": message_sid, "channel": "whatsapp"},
                auth=(runtime_twilio.account_sid, runtime_twilio.auth_token),
                headers={"Accept": "application/json"},
            )
        if 200 <= response.status_code < 300:
            log_event(logger, "whatsapp.read_indicator_sent", message_sid=message_sid)
            return True
        logger.info(
            "whatsapp.read_indicator_not_sent",
            extra={
                "message_sid": message_sid,
                "status_code": response.status_code,
                "response_preview": response.text[:300],
            },
        )
        return False
    except Exception as exc:  # pragma: no cover - red/servicio externo
        logger.info(
            "whatsapp.read_indicator_failed",
            extra={"message_sid": message_sid, "error": str(exc)},
        )
        return False


async def send_manual_message(
    *,
    to_number: str,
    body: str | None = None,
    template_sid: str | None = None,
    template_variables: dict[str, str] | None = None,
    template_name: str | None = None,
    template_language: str | None = None,
    attachments: list[dict[str, Any]] | None = None,
    organizacion_id: UUID | str | None = None,
) -> TwilioSendResult:
    """Expone el envío de mensajes manuales desde el panel o automatizaciones."""
    org_uuid: UUID | None
    if isinstance(organizacion_id, UUID):
        org_uuid = organizacion_id
    else:
        org_uuid = _parse_org_uuid(organizacion_id if isinstance(organizacion_id, str) else None)
    return await _send_whatsapp_reply(
        to_number=to_number,
        body=body,
        content_sid=template_sid,
        content_variables=template_variables,
        template_name=template_name,
        template_language=template_language,
        attachments=attachments,
        organizacion_id=org_uuid,
    )


def _build_openai_input(
    message: schemas.WhatsAppIncomingMessage,
    *,
    context_data: dict[str, Any] | None = None,
    summary_text: str | None = None,
    summary_created_en: str | None = None,
    catalog_context: str | None = None,
    attachment_content_items: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Normaliza el contenido del mensaje con contexto CRM para la Responses API."""
    text_parts: list[str] = []
    if message.body:
        text_parts.append(message.body)
    if message.media:
        attachment_lines = [
            f"- ({item.index + 1}) {item.content_type or 'archivo'}: {item.url}"
            for item in message.media
        ]
        text_parts.append("El usuario adjuntó archivos:\n" + "\n".join(attachment_lines))
    if not text_parts:
        text_parts.append("(mensaje sin texto)")

    context_lines = build_crm_context_lines(context_data)
    if context_lines:
        text_parts.append("")
        text_parts.extend(context_lines)

    if summary_text:
        summary_header = "Resumen previo"
        if summary_created_en:
            summary_header += f" ({summary_created_en})"
        text_parts.append("")
        text_parts.append(f"{summary_header}: {summary_text}")

    user_message = {
        "role": "user",
        "content": [
            {
                "type": "input_text",
                "text": "\n\n".join(text_parts),
            }
        ],
    }
    if attachment_content_items:
        user_message["content"].extend(attachment_content_items)
    messages: list[dict[str, Any]] = []
    if catalog_context:
        messages.append(
            {
                "role": "developer",
                "content": [
                    {
                        "type": "input_text",
                        "text": catalog_context,
                    }
                ],
            }
        )
    messages.append(user_message)
    return messages


def _extract_text_from_response(payload: dict[str, Any]) -> str | None:
    output_items = payload.get("output") or []
    fragments: list[str] = []
    for item in output_items:
        if item.get("type") != "message":
            continue
        for content in item.get("content") or []:
            if content.get("type") == "output_text":
                text = content.get("text")
                if text:
                    fragments.append(str(text))
    if fragments:
        return "\n".join(fragment.strip() for fragment in fragments if fragment)
    if payload.get("status") == "completed" and payload.get("output"):
        return None
    if payload.get("status") == "requires_action":
        logger.warning("whatsapp.tool_call_unhandled", extra={"output": payload.get("output")})
    return None


def _ensure_metadata_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return dict(value)
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            return {}
    return {}


def _parse_interactive_payload(raw_payload: dict[str, Any]) -> dict[str, Any] | None:
    data_text = raw_payload.get("InteractiveData") or raw_payload.get("interactivedata")
    if not data_text:
        return None
    try:
        parsed = json.loads(data_text)
    except json.JSONDecodeError:
        return None
    if isinstance(parsed, dict):
        return parsed
    return None


def _extract_quick_reply_data(raw_payload: dict[str, Any]) -> dict[str, Any] | None:
    if not raw_payload:
        return None
    button_payload = raw_payload.get("ButtonPayload") or raw_payload.get("buttonpayload")
    button_text = raw_payload.get("ButtonText") or raw_payload.get("buttontext")
    interactive = _parse_interactive_payload(raw_payload)
    if not button_payload and interactive:
        reply_data = interactive.get("button_reply") or {}
        button_payload = reply_data.get("id") or button_payload
        button_text = button_text or reply_data.get("title") or reply_data.get("text")
    if not button_payload and not interactive:
        return None
    raw_fields = {
        "ButtonText": raw_payload.get("ButtonText"),
        "ButtonPayload": raw_payload.get("ButtonPayload"),
        "InteractiveData": raw_payload.get("InteractiveData"),
    }
    return {
        "payload": button_payload,
        "text": button_text or raw_payload.get("Body"),
        "interactive": interactive,
        "raw_fields": raw_fields,
    }


async def _maybe_handle_sales_acknowledgement(
    message: schemas.WhatsAppIncomingMessage,
) -> bool:
    """Detecta respuestas de botones del vendedor y marca la asignación como aceptada."""
    quick_reply = _extract_quick_reply_data(message.raw_payload or {})
    if not quick_reply:
        return False

    normalized_from = _normalize_phone_number(message.from_number)
    if not normalized_from:
        logger.warning("whatsapp.sales_ack.invalid_from_number")
        return True

    repo = CRMRepository()
    try:
        seller = await repo.find_sales_rep_by_phone(phone_e164=normalized_from)
    except CRMRepositoryError as exc:
        logger.warning(
            "whatsapp.sales_ack.sales_rep_lookup_failed",
            extra={"error": str(exc)},
        )
        return True

    if not seller:
        logger.info(
            "whatsapp.sales_ack.sales_rep_missing",
            extra={"from_number": normalized_from},
        )
        return True

    vendedor_id = seller.get("usuario_id")
    organizacion_ids = seller.get("organizacion_ids") or []
    try:
        pending = await repo.find_pending_sales_assignment(
            vendedor_id=vendedor_id,
            organizacion_ids=organizacion_ids,
        )
    except CRMRepositoryError as exc:
        logger.warning(
            "whatsapp.sales_ack.assignment_lookup_failed",
            extra={"error": str(exc)},
        )
        return True

    if not pending:
        logger.info(
            "whatsapp.sales_ack.no_pending_assignment",
            extra={"seller_id": str(vendedor_id)},
        )
        return True

    try:
        assignment_id = UUID(str(pending.get("id")))
    except (TypeError, ValueError):
        logger.warning(
            "whatsapp.sales_ack.invalid_assignment_id",
            extra={"assignment_id": pending.get("id")},
        )
        return True

    metadata = _ensure_metadata_dict(pending.get("metadata"))
    metadata["acknowledgement"] = {
        "type": "whatsapp_quick_reply",
        "button_text": quick_reply.get("text"),
        "button_payload": quick_reply.get("payload"),
        "message_sid": message.message_sid,
        "raw_fields": quick_reply.get("raw_fields"),
    }
    ack_time = datetime.now(timezone.utc)
    try:
        await repo.update_sales_assignment_ack(
            assignment_id=assignment_id,
            ack_user_id=vendedor_id,
            ack_time=ack_time,
            ack_via="whatsapp_quick_reply",
            metadata=metadata,
        )
    except CRMRepositoryError as exc:
        logger.warning(
            "whatsapp.sales_ack.update_failed",
            extra={"assignment_id": str(assignment_id), "error": str(exc)},
        )
        return True

    log_event(
        logger,
        "whatsapp.sales_acknowledged",
        assignment_id=str(assignment_id),
        vendedor_id=str(vendedor_id),
        button_payload=str(quick_reply.get("payload") or ""),
    )
    return True


def _normalize_phone_number(value: str | None) -> str | None:
    if not value:
        return None
    text = value.strip()
    if text.lower().startswith("whatsapp:"):
        text = text.split(":", 1)[1]
    return normalize_phone(text) or text


def _phone_lookup_candidates(value: str | None) -> list[str]:
    normalized = _normalize_phone_number(value)
    if not normalized:
        return []
    candidates: list[str] = []

    def _push(raw: str | None) -> None:
        text = _trim_text(raw)
        if not text:
            return
        if text not in candidates:
            candidates.append(text)

    _push(normalized)
    digits = re.sub(r"\D", "", normalized)
    if digits:
        _push(digits)
    if len(digits) >= 10:
        tail10 = digits[-10:]
        _push(tail10)
        _push("+52" + tail10)
        _push("+521" + tail10)
        _push("52" + tail10)
        _push("521" + tail10)
    if normalized.startswith("+521") and len(normalized) > 4:
        _push("+52" + normalized[4:])
    if normalized.startswith("+52") and not normalized.startswith("+521") and len(normalized) > 3:
        _push("+521" + normalized[3:])
    return candidates


def _build_prospeccion_inbox_context_patch(
    envio: dict[str, Any] | None,
) -> dict[str, Any]:
    patch: dict[str, Any] = {"source": "prospeccion"}
    if not isinstance(envio, dict):
        return patch

    payload = envio.get("payload") if isinstance(envio.get("payload"), dict) else {}
    payload_meta = payload.get("metadata") if isinstance(payload.get("metadata"), dict) else {}
    detalle = envio.get("detalle") if isinstance(envio.get("detalle"), dict) else {}

    batch_id = _trim_text(envio.get("batch_id"))
    campana_id = _trim_text(envio.get("campana_id")) or _trim_text(payload_meta.get("campana_id"))
    template_id = _trim_text(
        payload.get("template_id")
        or payload_meta.get("template_id")
        or payload_meta.get("template_id_snapshot")
    )
    template_slug = _trim_text(
        payload.get("template_slug")
        or payload_meta.get("template_slug")
        or payload_meta.get("template_slug_snapshot")
        or payload_meta.get("kw")
        or payload_meta.get("twilio_content_sid")
        or detalle.get("template_sid")
        or payload_meta.get("template_sid_snapshot")
    )
    template_label = _trim_text(
        payload.get("template_label")
        or payload_meta.get("template_label")
        or payload_meta.get("template_nombre")
        or payload_meta.get("template_name")
        or payload_meta.get("template_nombre_snapshot")
        or payload_meta.get("template_name_snapshot")
        or payload_meta.get("template_label_snapshot")
    )

    if batch_id:
        patch["batch_id"] = batch_id
    if campana_id:
        patch["campana_id"] = campana_id
    if template_id:
        patch["template_id"] = template_id
    if template_slug:
        patch["template_slug"] = template_slug.lower()
    if template_label:
        patch["template_label"] = template_label
    return patch


def _parse_iso_datetime(value: Any) -> datetime | None:
    if not value:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    normalized = raw.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed


async def _coalesce_inbound_burst(
    *,
    conversation_id: str,
    current_message_id: str | None,
    fallback_body: str | None,
) -> tuple[bool, str | None, dict[str, float]]:
    """Agrupa mensajes entrantes consecutivos y evita respuestas duplicadas."""

    debug_timings: dict[str, float] = {}
    message_id = str(current_message_id or "").strip()
    fallback_text = str(fallback_body or "").strip() or None
    if not conversation_id or not message_id:
        return True, fallback_text, debug_timings

    should_wait_for_more = _should_wait_for_inbound_burst(fallback_text)
    debug_timings["debounce_applied"] = 1.0 if should_wait_for_more else 0.0
    if should_wait_for_more and _WHATSAPP_INBOUND_DEBOUNCE_SECONDS > 0:
        debounce_started = time.perf_counter()
        await asyncio.sleep(_WHATSAPP_INBOUND_DEBOUNCE_SECONDS)
        debug_timings["debounce_sleep_ms"] = round((time.perf_counter() - debounce_started) * 1000, 2)

    try:
        fetch_started = time.perf_counter()
        recent_messages = await storage.fetch_recent_messages(
            conversation_id=conversation_id,
            limit=max(8, _WHATSAPP_INBOUND_MERGE_MAX_MESSAGES + 4),
        )
        debug_timings["fetch_recent_messages_ms"] = round((time.perf_counter() - fetch_started) * 1000, 2)
    except StorageError as exc:
        logger.warning(
            "whatsapp.inbound_burst_fetch_failed",
            extra={"conversation_id": conversation_id, "error": str(exc)},
        )
        return True, fallback_text, debug_timings

    if not recent_messages:
        return True, fallback_text, debug_timings

    latest_inbound: dict[str, Any] | None = None
    for row in reversed(recent_messages):
        if str(row.get("direccion") or "").strip().lower() == "entrante":
            latest_inbound = row
            break

    if not isinstance(latest_inbound, dict):
        return True, fallback_text, debug_timings

    latest_inbound_id = str(latest_inbound.get("id") or "").strip()
    if not latest_inbound_id:
        return True, fallback_text, debug_timings

    if latest_inbound_id != message_id:
        log_event(
            logger,
            "whatsapp.inbound_burst_skip_older_fragment",
            conversation_id=conversation_id,
            message_id=message_id,
            latest_inbound_id=latest_inbound_id,
        )
        return False, None, debug_timings

    latest_created = _parse_iso_datetime(latest_inbound.get("creado_en"))
    fragments: list[str] = []
    for row in reversed(recent_messages):
        if str(row.get("direccion") or "").strip().lower() != "entrante":
            if fragments:
                break
            continue
        if len(fragments) >= _WHATSAPP_INBOUND_MERGE_MAX_MESSAGES:
            break
        if latest_created is not None:
            created_at = _parse_iso_datetime(row.get("creado_en"))
            if created_at is None:
                if fragments:
                    break
            elif (latest_created - created_at).total_seconds() > _WHATSAPP_INBOUND_MERGE_MAX_WINDOW_SECONDS:
                break
        candidate = str(row.get("texto") or "").strip()
        if candidate:
            fragments.append(candidate)

    if not fragments:
        return True, fallback_text, debug_timings

    fragments.reverse()
    merged_text = " ".join(fragment for fragment in fragments if fragment).strip()
    if not merged_text:
        return True, fallback_text, debug_timings

    if len(fragments) > 1:
        log_event(
            logger,
            "whatsapp.inbound_burst_coalesced",
            conversation_id=conversation_id,
            message_id=message_id,
            fragments=len(fragments),
        )
    debug_timings["fragments_count"] = float(len(fragments))
    return True, merged_text, debug_timings


def _map_status_to_event(status: str | None) -> str | None:
    if not status:
        return None
    normalized = status.strip().lower()
    mapping = {
        "queued": "en_cola",
        "accepted": "en_cola",
        "sending": "enviado",
        "sent": "enviado",
        "delivered": "entregado",
        "read": "leido",
        "failed": "fallido",
        "undelivered": "fallido",
    }
    return mapping.get(normalized)


def _parse_org_uuid(value: str | None) -> UUID | None:
    if not value:
        return None
    try:
        return UUID(value)
    except (TypeError, ValueError):
        return None


def _map_status_to_envio_estado(status: str | None) -> str | None:
    event = _map_status_to_event(status)
    if not event:
        return None
    if event == "en_cola":
        return "enviado"
    return event


def _build_assistant_from_runtime(
    settings_values: tenant_runtime.WhatsappRuntimeSettings,
    *,
    prospeccion_mode: bool = False,
) -> AssistantConfig:
    if prospeccion_mode and settings_values.prospeccion_prompt_id:
        return AssistantConfig(
            assistant_id=None,
            prompt_id=settings_values.prospeccion_prompt_id,
            prompt_version=settings_values.prospeccion_prompt_version or settings_values.prompt_version,
            project_id=settings_values.project_id or settings.openai_project_id,
        )
    if settings_values.prompt_id:
        return AssistantConfig(
            assistant_id=None,
            prompt_id=settings_values.prompt_id,
            prompt_version=settings_values.prompt_version,
            project_id=settings_values.project_id or settings.openai_project_id,
        )
    target_id = settings_values.assistant_id or settings.openai_assistant_id
    if not target_id:
        raise RuntimeError("No se configuró un ASSISTANT_ID para WhatsApp")
    return AssistantConfig(
        assistant_id=target_id,
        project_id=settings_values.project_id or settings.openai_project_id,
    )
